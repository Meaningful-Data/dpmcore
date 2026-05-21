"""Parametrized syntax tests for all expressions in the validations xlsx.

Skipped automatically when tests/fixtures/validations_export.xlsx is absent.
"""

from pathlib import Path

import openpyxl
import pytest

from dpmcore.services.syntax import SyntaxService

_XLSX = (
    Path(__file__).resolve().parents[3]
    / "tests"
    / "fixtures"
    / "validations_export.xlsx"
)


def _load_params():
    if not _XLSX.exists():
        return [
            pytest.param(
                "",
                "",
                "",
                marks=pytest.mark.skip(reason=f"xlsx not found: {_XLSX}"),
            )
        ]
    wb = openpyxl.load_workbook(_XLSX)
    ws = wb["Validations"]
    headers = [cell.value for cell in ws[1]]
    params = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        code = str(d.get("Code") or "")
        for col in ("Expression", "Precondition"):
            val = d.get(col)
            if val and str(val).strip():
                params.append(
                    pytest.param(
                        code,
                        col,
                        str(val).strip(),
                        id=f"{code}-{col}",
                    )
                )
    return params


@pytest.fixture(scope="module")
def syntax_service():
    return SyntaxService()


@pytest.mark.parametrize("code,col,expression", _load_params())
def test_syntax(code, col, expression, syntax_service):
    result = syntax_service.validate(expression)
    assert result.is_valid, f"{code} | {col} | {result.error_message}"
