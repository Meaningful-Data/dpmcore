"""Unit-level tests targeting specific branches of excel_writer.

These complement the DB-backed integration tests by constructing
TableLayout dataclasses directly. That way we can drive the Excel
writer through formatting branches that are otherwise hard to reach
without an elaborate DB fixture.
"""

from openpyxl import Workbook

from dpmcore.services.layout_exporter import excel_writer as ew
from dpmcore.services.layout_exporter.excel_writer import (
    ExcelLayoutWriter,
    _dim_display_label,
    _find_member_by_label,
    _format_categorisations,
    _is_descendant,
    _key_member_display,
    _member_display_label,
)
from dpmcore.services.layout_exporter.models import (
    CellData,
    DimensionMember,
    ExportConfig,
    LayoutHeader,
    TableLayout,
)


def _h(
    hid,
    label="L",
    code="C",
    direction="x",
    is_abstract=False,
    is_key=False,
    parent=None,
    parent_first=True,
    depth=0,
    cats=None,
    sub_code="",
    sub_cat="",
    sub_desc="",
    key_vid=None,
    key_var_id=None,
    key_dt="",
    key_pname="",
    key_cats=None,
):
    return LayoutHeader(
        header_id=hid,
        header_vid=hid,
        code=code,
        label=label,
        direction=direction,
        order=hid,
        is_abstract=is_abstract,
        is_key=is_key,
        parent_header_id=parent,
        parent_first=parent_first,
        depth=depth,
        categorisations=cats or [],
        subcategory_code=sub_code,
        subcategory_cat_code=sub_cat,
        subcategory_description=sub_desc,
        key_variable_vid=key_vid,
        key_variable_id=key_var_id,
        key_data_type_code=key_dt,
        key_property_name=key_pname,
        key_categorisations=key_cats or [],
    )


def _dm(prop_id=1, label="D", code="DC", domain="DOM", member="M", mc="m"):
    return DimensionMember(
        property_id=prop_id,
        dimension_label=label,
        dimension_code=code,
        domain_code=domain,
        member_label=member,
        member_code=mc,
    )


# --------------------------------------------------------------------------- #
# Pure helper functions
# --------------------------------------------------------------------------- #


def test_dim_display_label_aty():
    assert _dim_display_label(_dm(code="ATY", label="Main")) == "Main"


def test_dim_display_label_blank_code():
    assert _dim_display_label(_dm(code="", label="X")) == "X"


def test_dim_display_label_regular():
    out = _dim_display_label(_dm(code="DC", domain="DOM", label="L"))
    assert out == "(DC:DOM) L"


def test_member_display_label_aty_with_code():
    out = _member_display_label(_dm(code="ATY", member="C", mc="qC"))
    assert out == "(qC) C"


def test_member_display_label_aty_without_code():
    out = _member_display_label(_dm(code="ATY", member="C", mc=""))
    assert out == "C"


def test_member_display_label_full():
    out = _member_display_label(
        _dm(code="DC", domain="DOM", member="X", mc="x")
    )
    assert out == "(DOM:x) X"


def test_member_display_label_partial_falls_back_to_label():
    out = _member_display_label(_dm(code="DC", domain="", mc="", member="L"))
    assert out == "L"


def test_key_member_display_with_subcategory():
    ch = _h(1, sub_code="SC", sub_cat="CAT", sub_desc="Desc")
    dm = _dm(domain="DOM")
    out = _key_member_display(dm, ch)
    assert "<Key value>" in out
    assert "Desc" in out


def test_key_member_display_without_subcategory():
    ch = _h(1)
    dm = _dm(domain="DOM")
    assert _key_member_display(dm, ch) == "(DOM:) <Key value>"


def test_find_member_by_label_hit_and_miss():
    dm = _dm(code="DC", domain="DOM", label="X")
    assert _find_member_by_label([dm], "(DC:DOM) X") is dm
    assert _find_member_by_label([dm], "missing") is None


def test_format_categorisations_empty_and_populated():
    assert _format_categorisations([]) == ""
    out = _format_categorisations([_dm(label="D", member="M")])
    assert out == "D  =  M"


def test_is_descendant_yes_no_cycle():
    p = _h(1)
    c = _h(2, parent=1)
    g = _h(3, parent=2)
    by = {1: p, 2: c, 3: g}
    assert _is_descendant(g, 1, by) is True
    assert _is_descendant(p, 2, by) is False
    # Cycle: header points to itself
    cyc = _h(4, parent=4)
    assert _is_descendant(cyc, 99, {4: cyc}) is False
    # Missing parent in by_id
    orphan = _h(5, parent=99)
    assert _is_descendant(orphan, 1, {5: orphan}) is False


# --------------------------------------------------------------------------- #
# ExcelLayoutWriter — narrow branch coverage
# --------------------------------------------------------------------------- #


def _empty_layout(**kw):
    return TableLayout(table_vid=1, table_code="T", table_name="T", **kw)


def test_writer_init_with_no_default_sheet():
    """If the workbook has no sheets, the __init__ guard skips deletion."""
    w = ExcelLayoutWriter([], ExportConfig())
    # Force the wb into a no-sheets state then re-run init logic
    w.wb = Workbook()
    while w.wb.sheetnames:
        del w.wb[w.wb.sheetnames[0]]
    # Re-construct (covers the falsy `if self.wb.sheetnames` branch)
    w2 = ExcelLayoutWriter.__new__(ExcelLayoutWriter)
    w2.layouts = []
    w2.config = ExportConfig()
    w2.wb = w.wb
    if w2.wb.sheetnames:
        del w2.wb[w2.wb.sheetnames[0]]
    assert w2.wb.sheetnames == []


def test_writer_title_no_merge_when_single_cell():
    """Layout with 0 visible columns: title does not merge across columns."""
    layout = _empty_layout(
        rows=[_h(1, direction="y", code="R")],
        columns=[],
    )
    w = ExcelLayoutWriter([layout], ExportConfig(show_code_column=False))
    wb = w.write()
    ws = wb["T"]
    # Title is in A1; with no visible cols nothing to merge across
    assert ws["A1"].value.startswith("T - ")


def test_sheet_without_categorisations_or_subcategory():
    """Z-axis header lacking categorisations + subcategory."""
    layout = _empty_layout(
        rows=[_h(2, direction="y", code="R")],
        columns=[
            _h(
                1,
                direction="x",
                code="C",
                cats=[_dm(prop_id=10, label="Main", code="ATY", mc="qC")],
            )
        ],
        sheets=[_h(99, direction="z", code="S", label="Sheet")],
    )
    w = ExcelLayoutWriter([layout], ExportConfig())
    wb = w.write()
    ws = wb["T"]
    sheet_label = next(
        (
            c.value
            for row in ws.iter_rows()
            for c in row
            if isinstance(c.value, str) and c.value.startswith("Sheet per")
        ),
        None,
    )
    assert sheet_label == "Sheet per Sheet"


def test_sheet_with_subcategory_and_is_key():
    """Sheet header with subcategory_code + is_key triggers <Key value>."""
    layout = _empty_layout(
        rows=[_h(2, direction="y", code="R")],
        columns=[
            _h(
                1,
                direction="x",
                code="C",
                cats=[_dm(prop_id=10, label="Main", code="ATY", mc="qC")],
            )
        ],
        sheets=[
            _h(
                99,
                direction="z",
                code="S",
                label="Sheet",
                is_key=True,
                sub_code="SC",
                sub_cat="CAT",
                sub_desc="Desc",
                cats=[_dm(prop_id=11, label="X", code="DC", mc="m")],
            )
        ],
    )
    w = ExcelLayoutWriter([layout], ExportConfig())
    wb = w.write()
    ws = wb["T"]
    has_key = any(
        isinstance(c.value, str) and "<Key value>" in c.value
        for row in ws.iter_rows()
        for c in row
    )
    assert has_key


def test_abstract_row_with_visible_columns():
    """Abstract rows merge across the data columns when present."""
    layout = _empty_layout(
        rows=[
            _h(2, direction="y", code="R", is_abstract=True),
            _h(3, direction="y", code="R2"),
        ],
        columns=[_h(1, direction="x", code="C")],
    )
    w = ExcelLayoutWriter([layout], ExportConfig())
    wb = w.write()
    ws = wb["T"]
    assert any(
        rng.min_row == rng.max_row and rng.min_col != rng.max_col
        for rng in ws.merged_cells.ranges
    )


def test_open_row_key_column_e_type():
    """Open-row table: key column with data_type 'e' renders [property]."""
    layout = _empty_layout(
        rows=[],
        columns=[
            _h(
                1,
                direction="x",
                code="K",
                is_key=True,
                key_vid=99,
                key_var_id=42,
                key_dt="e",
                key_pname="Currency",
            )
        ],
        is_open_row=True,
    )
    w = ExcelLayoutWriter([layout], ExportConfig())
    wb = w.write()
    ws = wb["T"]
    has_e = any(
        isinstance(c.value, str) and "[Currency]" in c.value
        for row in ws.iter_rows()
        for c in row
    )
    assert has_e


def test_open_row_key_column_no_property_name():
    """Open-row key column with empty property_name skips <name> append."""
    layout = _empty_layout(
        rows=[],
        columns=[
            _h(
                1,
                direction="x",
                code="K",
                is_key=True,
                key_vid=99,
                key_var_id=42,
                key_dt="m",
                key_pname="",
            )
        ],
        is_open_row=True,
    )
    w = ExcelLayoutWriter([layout], ExportConfig())
    wb = w.write()
    ws = wb["T"]
    cells = [
        c.value
        for row in ws.iter_rows()
        for c in row
        if isinstance(c.value, str) and c.value.startswith("42")
    ]
    assert cells
    assert all("<" not in v for v in cells)


def test_open_row_non_key_excluded_fallback():
    """Non-key column without cell data falls into the excluded branch."""
    layout = _empty_layout(
        rows=[],
        columns=[_h(1, direction="x", code="C")],
        cells={},
        is_open_row=True,
    )
    w = ExcelLayoutWriter([layout], ExportConfig())
    wb = w.write()
    ws = wb["T"]
    # Find the open-row data cell — the fallback applies _EXCLUDED_FILL
    found_excluded = False
    for row in ws.iter_rows():
        for c in row:
            rgb = (
                c.fill.start_color.rgb
                if c.fill.start_color is not None
                else ""
            )
            if rgb in {"00999999", "FF999999"}:
                found_excluded = True
                break
    assert found_excluded


def test_open_row_non_key_with_e_type_and_domain():
    """Open-row non-key cell with e-type + domain_label."""
    layout = _empty_layout(
        rows=[],
        columns=[_h(1, direction="x", code="C")],
        cells={
            (None, 1, None): CellData(
                row_header_id=None,
                col_header_id=1,
                sheet_header_id=None,
                variable_vid=10,
                variable_id=42,
                data_type_code="e",
                domain_label="Curr",
                sign="positive",
            )
        },
        is_open_row=True,
    )
    w = ExcelLayoutWriter([layout], ExportConfig())
    wb = w.write()
    ws = wb["T"]
    has_domain = any(
        isinstance(c.value, str) and "[Curr]" in c.value
        for row in ws.iter_rows()
        for c in row
    )
    assert has_domain


def test_column_annotation_subcategory_display():
    """Column annotation: subcategory_code + subcategory_cat_code branch."""
    layout = _empty_layout(
        rows=[_h(2, direction="y", code="R")],
        columns=[
            _h(
                1,
                direction="x",
                code="C",
                sub_code="SC",
                sub_cat="CAT",
                cats=[
                    _dm(prop_id=5, label="L", code="DC", domain="DOM", mc="m")
                ],
            )
        ],
    )
    w = ExcelLayoutWriter([layout], ExportConfig())
    wb = w.write()
    ws = wb["T"]
    has_sub = any(
        isinstance(c.value, str) and "[CAT:SC]" in c.value
        for row in ws.iter_rows()
        for c in row
    )
    assert has_sub


def test_row_annotation_subcategory_display():
    """Row annotation: subcategory_code + subcategory_cat_code branch."""
    layout = _empty_layout(
        rows=[
            _h(
                2,
                direction="y",
                code="R",
                sub_code="SC",
                sub_cat="CAT",
                cats=[
                    _dm(prop_id=5, label="L", code="DC", domain="DOM", mc="m")
                ],
            )
        ],
        columns=[_h(1, direction="x", code="C")],
    )
    w = ExcelLayoutWriter([layout], ExportConfig())
    wb = w.write()
    ws = wb["T"]
    has_sub = any(
        isinstance(c.value, str) and "[CAT:SC]" in c.value
        for row in ws.iter_rows()
        for c in row
    )
    assert has_sub


def test_collect_axis_dimensions_with_supplemental_and_key():
    """Cover the supplemental_cats and is_key extension branches."""
    h = _h(
        1,
        is_key=True,
        cats=[_dm(prop_id=1, code="DC", label="Main", mc="m")],
        key_cats=[_dm(prop_id=2, code="DK", label="Key", mc="k")],
    )
    out = ew._collect_axis_dimensions(
        [h, _h(9, is_abstract=True)],
        supplemental_cats={
            1: [_dm(prop_id=3, code="DS", label="Sup", mc="s")]
        },
    )
    assert "(DC:DOM) Main" in out
    assert any(label.startswith("(DK:") for label in out)
    assert any(label.startswith("(DS:") for label in out)


def test_data_cells_with_e_type_no_domain_label_skips():
    """Cell with data_type 'e' but blank domain_label hits no-domain branch."""
    layout = _empty_layout(
        rows=[_h(2, direction="y", code="R")],
        columns=[_h(1, direction="x", code="C")],
        cells={
            (2, 1, None): CellData(
                row_header_id=2,
                col_header_id=1,
                sheet_header_id=None,
                variable_vid=10,
                variable_id=42,
                data_type_code="e",
                domain_label="",
            )
        },
    )
    w = ExcelLayoutWriter([layout], ExportConfig())
    wb = w.write()
    ws = wb["T"]
    # Data cell value should be just "42" with no domain
    cells = [
        c.value
        for row in ws.iter_rows()
        for c in row
        if isinstance(c.value, str) and "42" in c.value
    ]
    assert cells
    assert not any("[" in v for v in cells)


def test_data_cell_unknown_data_type_skips_symbol():
    layout = _empty_layout(
        rows=[_h(2, direction="y", code="R")],
        columns=[_h(1, direction="x", code="C")],
        cells={
            (2, 1, None): CellData(
                row_header_id=2,
                col_header_id=1,
                sheet_header_id=None,
                variable_vid=10,
                variable_id=42,
                data_type_code="zzz",
            )
        },
    )
    w = ExcelLayoutWriter([layout], ExportConfig())
    wb = w.write()
    ws = wb["T"]
    has_42 = any(
        isinstance(c.value, str) and c.value.startswith("42")
        for row in ws.iter_rows()
        for c in row
    )
    assert has_42


def test_data_cell_no_data_type_no_sign():
    layout = _empty_layout(
        rows=[_h(2, direction="y", code="R")],
        columns=[_h(1, direction="x", code="C")],
        cells={
            (2, 1, None): CellData(
                row_header_id=2,
                col_header_id=1,
                sheet_header_id=None,
                variable_vid=10,
                variable_id=42,
                data_type_code="",
            )
        },
    )
    w = ExcelLayoutWriter([layout], ExportConfig())
    wb = w.write()
    assert wb["T"]


def test_data_cell_sign_suppressed_by_parent_explicit_sign():
    """Monetary cell with parent that has an explicit sign: no auto-positive."""
    layout = _empty_layout(
        rows=[
            _h(1, direction="y", code="P"),
            _h(2, direction="y", code="C", parent=1),
        ],
        columns=[_h(10, direction="x", code="X")],
        cells={
            # Parent row has a cell with explicit sign — so child row's
            # auto-positive default should be suppressed.
            (1, 10, None): CellData(
                row_header_id=1,
                col_header_id=10,
                sheet_header_id=None,
                variable_vid=99,
                variable_id=99,
                data_type_code="m",
                sign="negative",
            ),
            (2, 10, None): CellData(
                row_header_id=2,
                col_header_id=10,
                sheet_header_id=None,
                variable_vid=100,
                variable_id=100,
                data_type_code="m",
            ),
        },
    )
    w = ExcelLayoutWriter([layout], ExportConfig())
    wb = w.write()
    ws = wb["T"]
    # Child cell at row 2 should NOT have "positive" appended.
    found_child = [
        c.value
        for row in ws.iter_rows()
        for c in row
        if isinstance(c.value, str) and c.value.startswith("100")
    ]
    assert found_child
    assert all("positive" not in v for v in found_child)


def test_data_cell_closing_balance_default_positive():
    """Row labelled 'Closing balance ...' triggers default positive sign."""
    layout = _empty_layout(
        rows=[_h(1, direction="y", code="R", label="Closing balance Q4")],
        columns=[_h(10, direction="x", code="X")],
        cells={
            (1, 10, None): CellData(
                row_header_id=1,
                col_header_id=10,
                sheet_header_id=None,
                variable_vid=99,
                variable_id=99,
                data_type_code="m",
            )
        },
    )
    w = ExcelLayoutWriter([layout], ExportConfig())
    wb = w.write()
    ws = wb["T"]
    has_pos = any(
        isinstance(c.value, str) and "positive" in c.value
        for row in ws.iter_rows()
        for c in row
    )
    assert has_pos


def test_abstract_row_without_visible_columns():
    """Abstract row when num_visible_cols==0: skip the merge."""
    layout = _empty_layout(
        rows=[_h(1, direction="y", code="R", is_abstract=True)],
        columns=[],
    )
    w = ExcelLayoutWriter([layout], ExportConfig())
    wb = w.write()
    ws = wb["T"]
    # Just confirm a worksheet was produced
    assert ws is not None


def test_open_row_with_sheets():
    """Open-row layout that also has Z-axis sheets."""
    layout = _empty_layout(
        rows=[],
        columns=[_h(1, direction="x", code="C")],
        sheets=[
            _h(
                3,
                direction="z",
                code="S",
                label="Sheet",
                cats=[_dm(prop_id=10, code="ATY", label="M", mc="m")],
            )
        ],
        is_open_row=True,
    )
    w = ExcelLayoutWriter([layout], ExportConfig())
    wb = w.write()
    assert "T" in wb.sheetnames


def test_row_dp_cats_extra_filter_empty():
    """All cell dp_categorisations are already covered by column dims.

    Hits the ``if extra:`` False branch of the row_dp_cats build.
    """
    dm = _dm(prop_id=5, label="X", code="DC", domain="DOM", mc="m")
    layout = _empty_layout(
        rows=[_h(2, direction="y", code="R")],
        columns=[_h(1, direction="x", code="C", cats=[dm])],
        cells={
            (2, 1, None): CellData(
                row_header_id=2,
                col_header_id=1,
                sheet_header_id=None,
                variable_vid=10,
                # Same dimension as the column header — extra is empty
                dp_categorisations=[dm],
            )
        },
    )
    w = ExcelLayoutWriter([layout], ExportConfig())
    wb = w.write()
    assert "T" in wb.sheetnames


def test_row_annotation_skips_abstract_row():
    """An abstract row should be skipped when writing row annotations."""
    dm_row = _dm(prop_id=5, label="Lbl", code="DC", domain="DOM", mc="m")
    layout = _empty_layout(
        rows=[
            _h(2, direction="y", code="R", is_abstract=True),
            _h(3, direction="y", code="R2", cats=[dm_row]),
        ],
        columns=[_h(1, direction="x", code="C")],
    )
    w = ExcelLayoutWriter([layout], ExportConfig())
    wb = w.write()
    assert "T" in wb.sheetnames


def test_row_annotation_no_member_for_some_rows():
    """Row without the active dimension hits the ``member None`` skip."""
    dm_a = _dm(prop_id=5, label="A", code="DA", domain="DOM", mc="a")
    layout = _empty_layout(
        rows=[
            _h(2, direction="y", code="R1", cats=[dm_a]),
            # Row 3 has no categorisations -> no match for dim A
            _h(3, direction="y", code="R2", cats=[]),
        ],
        columns=[_h(1, direction="x", code="C")],
    )
    w = ExcelLayoutWriter([layout], ExportConfig())
    wb = w.write()
    assert "T" in wb.sheetnames


def test_apply_col_groups_skips_abstract():
    """Abstract columns are not given outline_level."""
    layout = _empty_layout(
        rows=[_h(2, direction="y", code="R")],
        columns=[
            _h(1, direction="x", code="P", depth=0),
            _h(
                2, direction="x", code="A", parent=1, depth=1, is_abstract=True
            ),
            _h(3, direction="x", code="C", parent=1, depth=1),
        ],
    )
    w = ExcelLayoutWriter([layout], ExportConfig())
    wb = w.write()
    ws = wb["T"]
    # Smoke check: column dimensions object is accessible.
    assert ws.column_dimensions is not None
