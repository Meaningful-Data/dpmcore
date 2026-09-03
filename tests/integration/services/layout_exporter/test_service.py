"""Integration tests for LayoutExporterService."""

from __future__ import annotations

import os
from pathlib import Path

import pytest
from _helpers import (  # noqa: E402  (sys.path injected via conftest)
    add_cell,
    add_header,
    add_item_category,
    add_subcategory,
    add_subcategory_item,
    add_table,
    add_variable_version,
    build_basic_module_with_table,
    make_member,
    make_module,
    make_property,
    seed_data_types,
    seed_domain_category,
    seed_property_category,
    seed_releases,
)
from openpyxl import load_workbook

from dpmcore.orm.glossary import ItemCategory
from dpmcore.orm.packaging import ModuleVersion, ModuleVersionComposition
from dpmcore.orm.rendering import TableVersion
from dpmcore.services.layout_exporter.service import (
    LayoutExporterService,
    _fix_xlsx_timestamps,
)


def test_service_init_stores_session(memory_session):
    svc = LayoutExporterService(memory_session)
    assert svc.session is memory_session


def test_export_module_writes_workbook(memory_session, tmp_path):
    build_basic_module_with_table(memory_session)
    svc = LayoutExporterService(memory_session)
    out = svc.export_module("MOD1", output_path=str(tmp_path / "x.xlsx"))
    assert out.exists()
    wb = load_workbook(out)
    assert "T1" in wb.sheetnames
    assert "Index" in wb.sheetnames


def test_export_module_default_output_path(memory_session, tmp_path):
    build_basic_module_with_table(memory_session)
    svc = LayoutExporterService(memory_session)
    cwd = Path.cwd()
    try:
        os.chdir(tmp_path)
        out = svc.export_module("MOD1")
        assert out == Path("MOD1.xlsx")
        assert out.exists()
    finally:
        os.chdir(cwd)


def test_export_module_raises_when_no_tables(memory_session):
    svc = LayoutExporterService(memory_session)
    with pytest.raises(ValueError, match="No tables found"):
        svc.export_module("UNKNOWN")


def test_export_module_raises_includes_release_in_msg(memory_session):
    """When tables are missing for a known release, the message names it.

    ``release_code`` is a range filter (resolves through
    ``resolve_release_id`` + ``filter_by_release``), so the code must
    refer to a real ``Release`` row. An *unknown* code raises a
    different error from ``resolve_release_id`` itself; that case is
    covered separately in :func:`test_export_module_unknown_release_raises`.
    """
    seed_releases(memory_session)  # adds 1.0 and 2.0
    memory_session.commit()
    svc = LayoutExporterService(memory_session)
    with pytest.raises(
        ValueError,
        match=r"No tables found.*at release '1.0'",
    ):
        svc.export_module("UNKNOWN", release_code="1.0")


def test_export_module_unknown_release_raises(memory_session):
    """An unknown release code surfaces the resolver's ValueError."""
    svc = LayoutExporterService(memory_session)
    with pytest.raises(ValueError, match="not found"):
        svc.export_module("UNKNOWN", release_code="999.0")


def test_export_module_filters_empty_layouts(memory_session, tmp_path):
    """A table with no rows/cols should not appear in the workbook."""
    seed_releases(memory_session)
    seed_data_types(memory_session)
    seed_property_category(memory_session)
    make_module(memory_session, module_id=1, module_vid=10, code="MOD1")
    # Two tables: the second has no headers and should be filtered out
    add_table(
        memory_session,
        table_id=100,
        table_vid=1000,
        code="T_OK",
        name="OK",
        module_vid=10,
        order=1,
    )
    add_table(
        memory_session,
        table_id=101,
        table_vid=1001,
        code="T_EMPTY",
        name="Empty",
        module_vid=10,
        order=2,
    )
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=1,
        header_vid=11,
        direction="x",
        code="010",
        label="Col",
    )
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=2,
        header_vid=12,
        direction="y",
        code="010",
        label="Row",
    )
    memory_session.commit()

    svc = LayoutExporterService(memory_session)
    out = svc.export_module("MOD1", output_path=str(tmp_path / "out.xlsx"))
    wb = load_workbook(out)
    assert "T_OK" in wb.sheetnames
    assert "T_EMPTY" not in wb.sheetnames


def test_export_tables_writes_workbook(memory_session, tmp_path):
    build_basic_module_with_table(memory_session)
    svc = LayoutExporterService(memory_session)
    out = svc.export_tables(["T1"], output_path=str(tmp_path / "x.xlsx"))
    assert out.exists()


def test_export_tables_default_output_path(memory_session, tmp_path):
    build_basic_module_with_table(memory_session)
    svc = LayoutExporterService(memory_session)
    cwd = Path.cwd()
    try:
        os.chdir(tmp_path)
        out = svc.export_tables(["T1"])
        assert out == Path("tables.xlsx")
        assert out.exists()
    finally:
        os.chdir(cwd)


def test_export_tables_skips_missing_codes(memory_session, tmp_path):
    build_basic_module_with_table(memory_session)
    svc = LayoutExporterService(memory_session)
    out = svc.export_tables(
        ["T1", "MISSING"], output_path=str(tmp_path / "x.xlsx")
    )
    wb = load_workbook(out)
    assert "T1" in wb.sheetnames
    assert "MISSING" not in wb.sheetnames


def test_export_tables_raises_when_all_missing(memory_session):
    svc = LayoutExporterService(memory_session)
    with pytest.raises(ValueError, match="No valid tables"):
        svc.export_tables(["NOPE1", "NOPE2"])


def test_export_tables_skips_layout_without_rows_or_cols(
    memory_session, tmp_path
):
    """An existing table with no headers is filtered out of the workbook."""
    build_basic_module_with_table(memory_session)
    add_table(
        memory_session,
        table_id=200,
        table_vid=2000,
        code="EMPTY",
        name="Empty",
    )
    memory_session.commit()
    svc = LayoutExporterService(memory_session)
    out = svc.export_tables(
        ["EMPTY", "T1"],
        output_path=str(tmp_path / "x.xlsx"),
    )
    wb = load_workbook(out)
    assert "EMPTY" not in wb.sheetnames
    assert "T1" in wb.sheetnames


def test_build_layout_skips_cell_without_variable_vid(memory_session):
    """A cell without variable_vid still produces a valid layout."""
    seed_releases(memory_session)
    seed_data_types(memory_session)
    seed_property_category(memory_session)
    seed_domain_category(memory_session, 10, "DOM")
    make_module(memory_session, module_id=1, module_vid=10, code="MODX")
    add_table(
        memory_session,
        table_id=100,
        table_vid=1000,
        code="TX",
        name="X",
        module_vid=10,
    )
    make_property(
        memory_session,
        property_id=200,
        name="Carrying amount",
        data_type_id=1,
        dim_code="qC",
        domain_category_id=10,
    )
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=1,
        header_vid=11,
        direction="x",
        code="010",
        label="Col",
        property_id=200,
    )
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=2,
        header_vid=12,
        direction="y",
        code="010",
        label="Row",
    )
    # Cell without variable_vid (None)
    add_cell(
        memory_session,
        cell_id=900,
        table_id=100,
        table_vid=1000,
        column_id=1,
        row_id=2,
        variable_vid=None,
    )
    memory_session.commit()
    svc = LayoutExporterService(memory_session)
    layout = svc.build_layout("TX")
    assert (2, 1, None) in layout.cells


def test_build_layout_key_column_without_property(memory_session):
    """Key column whose key variable points to a missing property_id.

    Triggers the ``if atm_dm:`` False branch for key-column annotations
    (atm_dm is None when the property has no Item/Property row).
    """
    seed_releases(memory_session)
    seed_data_types(memory_session)
    seed_property_category(memory_session)
    make_module(memory_session, module_id=1, module_vid=10, code="MODK")
    add_table(
        memory_session,
        table_id=100,
        table_vid=1000,
        code="TK",
        name="K",
        module_vid=10,
    )
    # Variable references a property_id that has no Item/Property,
    # so load_property_as_categorisation returns nothing for it.
    add_variable_version(
        memory_session,
        variable_id=400,
        variable_vid=4000,
        code="V",
        property_id=999,
    )
    # Open-row table: only a key column, no rows
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=1,
        header_vid=11,
        direction="x",
        code="010",
        label="Key",
        is_key=True,
        key_variable_vid=4000,
    )
    memory_session.commit()
    svc = LayoutExporterService(memory_session)
    layout = svc.build_layout("TK")
    assert layout.columns
    assert layout.columns[0].key_categorisations == []


def test_build_layout_returns_layout(memory_session):
    build_basic_module_with_table(memory_session)
    svc = LayoutExporterService(memory_session)
    layout = svc.build_layout("T1")
    assert layout.table_code == "T1"
    assert len(layout.columns) == 1
    assert len(layout.rows) == 1


def test_build_layout_raises_when_table_missing(memory_session):
    svc = LayoutExporterService(memory_session)
    with pytest.raises(ValueError, match="not found"):
        svc.build_layout("NOPE")


def test_build_layout_with_release_code(memory_session):
    build_basic_module_with_table(memory_session)
    svc = LayoutExporterService(memory_session)
    layout = svc.build_layout("T1", release_code="1.0")
    assert layout.table_code == "T1"


def test_build_layout_full_pipeline_with_key_columns(memory_session):
    """Exercise key-variable annotation path in _build_layout."""
    seed_releases(memory_session)
    seed_data_types(memory_session)
    seed_property_category(memory_session)
    seed_domain_category(memory_session, 20, "DOMK")

    make_module(memory_session, module_id=1, module_vid=10, code="MOD1")
    add_table(
        memory_session,
        table_id=100,
        table_vid=1000,
        code="T_KEY",
        name="Key Table",
        module_vid=10,
    )

    # Property used by the key variable
    make_property(
        memory_session,
        property_id=200,
        name="Currency",
        data_type_id=2,  # 'e'
        dim_code="qCUR",
        domain_category_id=20,
    )
    add_variable_version(
        memory_session,
        variable_id=400,
        variable_vid=4000,
        code="VK",
        property_id=200,
    )

    # Open-row layout: column with key_variable_vid, no rows
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=1,
        header_vid=11,
        direction="x",
        code="010",
        label="Col",
        is_key=True,
        key_variable_vid=4000,
    )

    # Subcategory referenced by header (won't be used here)
    memory_session.commit()

    svc = LayoutExporterService(memory_session)
    layout = svc.build_layout("T_KEY")
    assert len(layout.columns) == 1
    col = layout.columns[0]
    assert col.is_key is True
    assert col.key_variable_id == 400
    assert col.key_data_type_code == "e"
    assert col.key_property_name == "Currency"
    # key_categorisations should be populated (synthetic from property dim)
    assert len(col.key_categorisations) == 1
    assert col.key_categorisations[0].dimension_label == "Currency"


def test_build_layout_with_subcategory_branch(memory_session):
    """Exercise subcategory_info branch in _build_layout."""
    seed_releases(memory_session)
    seed_data_types(memory_session)
    seed_property_category(memory_session)
    seed_domain_category(memory_session, 30, "BAS")

    make_module(memory_session, module_id=1, module_vid=10, code="MOD1")
    add_table(
        memory_session,
        table_id=100,
        table_vid=1000,
        code="T_SUB",
        name="Sub Table",
        module_vid=10,
    )

    add_subcategory(
        memory_session,
        subcategory_id=1,
        subcategory_vid=11,
        category_id=30,
        code="SC1",
        description="My Sub",
    )

    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=1,
        header_vid=11,
        direction="z",
        code="010",
        label="Sheet 1",
        subcategory_vid=11,
    )
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=2,
        header_vid=12,
        direction="x",
        code="010",
        label="Col",
    )
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=3,
        header_vid=13,
        direction="y",
        code="010",
        label="Row",
    )
    memory_session.commit()

    svc = LayoutExporterService(memory_session)
    layout = svc.build_layout("T_SUB")
    assert len(layout.sheets) == 1
    assert layout.sheets[0].subcategory_code == "SC1"
    assert layout.sheets[0].subcategory_description == "My Sub"
    assert layout.sheets[0].subcategory_cat_code == "BAS"


def test_build_layout_default_sheet_id_branch(memory_session, tmp_path):
    """When no Z headers but cells have sheet_id, that gets collected."""
    seed_releases(memory_session)
    seed_data_types(memory_session)
    seed_property_category(memory_session)
    make_module(memory_session, module_id=1, module_vid=10, code="MOD1")
    add_table(
        memory_session,
        table_id=100,
        table_vid=1000,
        code="T_DSH",
        name="Default Sheet",
        module_vid=10,
    )

    # Header used as the sheet ID (must exist as a Header row)
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=99,
        header_vid=99,
        direction="x",  # not z, so it's used as a default sheet id
        code="DSH",
        label="DSH",
    )
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=1,
        header_vid=11,
        direction="x",
        code="010",
        label="Col",
    )
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=2,
        header_vid=12,
        direction="y",
        code="010",
        label="Row",
    )
    add_variable_version(
        memory_session, variable_id=400, variable_vid=4000, code="V"
    )
    add_cell(
        memory_session,
        cell_id=900,
        table_id=100,
        table_vid=1000,
        column_id=1,
        row_id=2,
        sheet_id=99,  # cell has a sheet_id
        variable_vid=4000,
    )
    memory_session.commit()

    svc = LayoutExporterService(memory_session)
    layout = svc.build_layout("T_DSH")
    # The cell should be present using the cell's sheet_id
    assert (2, 1, 99) in layout.cells


# ---------------------------------------------------------------- #
# _fix_xlsx_timestamps
# ---------------------------------------------------------------- #


def test_fix_xlsx_timestamps_replaces_offset(tmp_path):
    """When core.xml contains '+00:00Z' it must be replaced with 'Z'."""
    import zipfile

    path = tmp_path / "weird.xlsx"
    with zipfile.ZipFile(path, "w") as zf:
        zf.writestr(
            "docProps/core.xml",
            "<root>2026-03-31T16:00:00+00:00Z</root>",
        )
        zf.writestr("other.xml", "data")
    _fix_xlsx_timestamps(path)
    with zipfile.ZipFile(path, "r") as zf:
        core = zf.read("docProps/core.xml").decode("utf-8")
    assert "+00:00Z" not in core
    assert "+00:00" not in core
    assert "Z</root>" in core


def test_fix_xlsx_timestamps_no_op_without_core_xml(tmp_path):
    """Workbook without docProps/core.xml is left untouched."""
    import zipfile

    path = tmp_path / "no_core.xlsx"
    with zipfile.ZipFile(path, "w") as zf:
        zf.writestr("other.xml", "x")
    raw = path.read_bytes()
    _fix_xlsx_timestamps(path)
    assert path.read_bytes() == raw


def test_fix_xlsx_timestamps_no_op_when_no_marker(tmp_path):
    """Core.xml without '+00:00Z' marker is left untouched."""
    import zipfile

    path = tmp_path / "ok.xlsx"
    with zipfile.ZipFile(path, "w") as zf:
        zf.writestr(
            "docProps/core.xml",
            "<root>2026-03-31T16:00:00Z</root>",
        )
    raw = path.read_bytes()
    _fix_xlsx_timestamps(path)
    assert path.read_bytes() == raw


def _build_enumerated_table(memory_session):
    """One enumerated cell whose column header carries the hierarchy."""
    seed_releases(memory_session)
    seed_data_types(memory_session)
    seed_property_category(memory_session)
    seed_domain_category(memory_session, 30, "DOM")

    make_module(memory_session, module_id=1, module_vid=10, code="MOD1")
    add_table(
        memory_session,
        table_id=100,
        table_vid=1000,
        code="T_ENUM",
        name="Enumerated Table",
        module_vid=10,
    )

    add_subcategory(
        memory_session,
        subcategory_id=1,
        subcategory_vid=11,
        category_id=30,
        code="SC1",
        name="Type of identifier",
    )
    make_member(
        memory_session,
        item_id=201,
        name="LEI code type",
        domain_category_id=30,
        code="m1",
        signature="eba_DOM:m1",
    )
    make_member(
        memory_session,
        item_id=202,
        name="MFI code",
        domain_category_id=30,
        code="m2",
        signature="eba_DOM:m2",
    )
    add_subcategory_item(
        memory_session, subcategory_vid=11, item_id=201, order=1
    )
    add_subcategory_item(
        memory_session, subcategory_vid=11, item_id=202, order=2
    )

    make_property(
        memory_session,
        property_id=200,
        name="Type of code",
        data_type_id=2,  # 'e'
        dim_code="qCO",
        domain_category_id=30,
    )
    add_variable_version(
        memory_session,
        variable_id=400,
        variable_vid=4000,
        code="VE",
        property_id=200,
    )
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=1,
        header_vid=11,
        direction="x",
        code="010",
        label="Col",
        subcategory_vid=11,
    )
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=2,
        header_vid=12,
        direction="y",
        code="010",
        label="Row",
    )
    add_cell(
        memory_session,
        cell_id=1,
        table_id=100,
        table_vid=1000,
        column_id=1,
        row_id=2,
        variable_vid=4000,
    )
    memory_session.commit()


def test_build_layout_attaches_possible_values_to_enumerated_cells(
    memory_session,
):
    _build_enumerated_table(memory_session)

    layout = LayoutExporterService(memory_session).build_layout("T_ENUM")

    cell = layout.cells[(2, 1, None)]
    assert cell.subcategory_vid == 11
    assert cell.enumeration is not None
    assert cell.enumeration.code == "SC1"
    assert [v.code for v in cell.enumeration.values] == ["m1", "m2"]


def test_export_writes_possible_values_as_a_cell_comment(
    memory_session, tmp_path
):
    _build_enumerated_table(memory_session)

    out = LayoutExporterService(memory_session).export_tables(
        ["T_ENUM"],
        output_path=str(tmp_path / "enum.xlsx"),
    )

    ws = load_workbook(out)["T_ENUM"]
    comments = [
        c.comment.text
        for row in ws.iter_rows()
        for c in row
        if c.comment and "Possible values" in c.comment.text
    ]
    assert len(comments) == 1
    assert "Possible values - (SC1) Type of identifier [2]:" in comments[0]
    assert "  (eba_DOM:m1) LEI code type" in comments[0]
    assert "  (eba_DOM:m2) MFI code" in comments[0]


def test_export_reads_codes_in_the_module_version_window(
    memory_session, tmp_path
):
    """A workbook shows the codes in force while its module ran.

    Member m1 is recoded at release 2.0, and module MOD1 has two
    versions either side of that: each must report its own code.
    """
    _build_enumerated_table(memory_session)
    # Close the seeded module version at 2.0 and add the next one,
    # composing the same table.
    memory_session.query(ModuleVersion).filter(
        ModuleVersion.module_vid == 10,
    ).update({"end_release_id": 2})
    memory_session.add(
        ModuleVersion(
            module_vid=20,
            module_id=1,
            code="MOD1",
            start_release_id=2,
        ),
    )
    memory_session.add(
        ModuleVersionComposition(
            module_vid=20,
            table_vid=1000,
            table_id=100,
            order=1,
        ),
    )
    # m1 is recoded to new1 at release 2.0.
    memory_session.query(ItemCategory).filter(
        ItemCategory.item_id == 201,
    ).update({"end_release_id": 2})
    add_item_category(
        memory_session,
        item_id=201,
        domain_category_id=30,
        code="new1",
        signature="eba_DOM:new1",
        start_release_id=2,
    )
    memory_session.commit()

    svc = LayoutExporterService(memory_session)

    def exported_values(release_code):
        out = svc.export_module(
            "MOD1",
            release_code,
            output_path=str(tmp_path / f"{release_code}.xlsx"),
        )
        ws = load_workbook(out)["T_ENUM"]
        return next(
            c.comment.text
            for row in ws.iter_rows()
            for c in row
            if c.comment and "Possible values" in c.comment.text
        )

    assert "(eba_DOM:m1) LEI code type" in exported_values("1.0")
    assert "(eba_DOM:new1) LEI code type" in exported_values("2.0")


def test_build_layout_reads_codes_in_the_table_version_window(memory_session):
    """Outside a module, the table version's own window is used."""
    _build_enumerated_table(memory_session)
    # The table version ran until 2.0; m1 was recoded at 2.0, which is
    # after that version was retired, so the old code must stand.
    memory_session.query(TableVersion).filter(
        TableVersion.table_vid == 1000,
    ).update({"end_release_id": 2})
    memory_session.query(ItemCategory).filter(
        ItemCategory.item_id == 201,
    ).update({"end_release_id": 2})
    add_item_category(
        memory_session,
        item_id=201,
        domain_category_id=30,
        code="new1",
        signature="eba_DOM:new1",
        start_release_id=2,
    )
    memory_session.commit()

    layout = LayoutExporterService(memory_session).build_layout(
        "T_ENUM",
        "1.0",
    )
    values = layout.cells[(2, 1, None)].enumeration.values
    assert values[0].signature == "eba_DOM:m1"


def test_build_layout_populates_key_fields_on_open_sheet_headers(
    memory_session,
):
    """An open-sheet table keys its sheets off the Z header's variable."""
    seed_releases(memory_session)
    seed_data_types(memory_session)
    seed_property_category(memory_session)
    seed_domain_category(memory_session, 30, "DOM")

    make_module(memory_session, module_id=1, module_vid=10, code="MOD1")
    add_table(
        memory_session,
        table_id=100,
        table_vid=1000,
        code="T_OPENZ",
        name="Open Sheet Table",
        module_vid=10,
    )
    add_subcategory(
        memory_session,
        subcategory_id=1,
        subcategory_vid=11,
        category_id=30,
        code="SC1",
        name="Exposure classes",
    )
    make_member(
        memory_session,
        item_id=201,
        name="Central governments",
        domain_category_id=30,
        code="m1",
        signature="eba_DOM:m1",
    )
    add_subcategory_item(
        memory_session, subcategory_vid=11, item_id=201, order=1
    )
    make_property(
        memory_session,
        property_id=200,
        name="Exposure class",
        data_type_id=2,  # 'e'
        dim_code="qEC",
        domain_category_id=30,
    )
    add_variable_version(
        memory_session,
        variable_id=400,
        variable_vid=4000,
        code="VZ",
        property_id=200,
    )
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=1,
        header_vid=11,
        direction="z",
        code="0010",
        label="Exposure class",
        is_key=True,
        key_variable_vid=4000,
        subcategory_vid=11,
    )
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=2,
        header_vid=12,
        direction="x",
        code="0010",
        label="Col",
    )
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=3,
        header_vid=13,
        direction="y",
        code="0010",
        label="Row",
    )
    memory_session.commit()

    layout = LayoutExporterService(memory_session).build_layout("T_OPENZ")

    sheet = layout.sheets[0]
    assert sheet.key_variable_id == 400
    assert sheet.key_data_type_code == "e"
    assert sheet.key_property_name == "Exposure class"
    assert sheet.key_enumeration is not None
    assert [v.signature for v in sheet.key_enumeration.values] == [
        "eba_DOM:m1",
    ]
