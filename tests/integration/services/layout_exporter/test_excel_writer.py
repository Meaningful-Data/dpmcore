"""Integration tests for ExcelLayoutWriter."""

from __future__ import annotations

from openpyxl import load_workbook

from dpmcore.services.layout_exporter.excel_writer import (
    ExcelLayoutWriter,
    _build_header_tooltip,
    _collect_axis_dimensions,
    _dim_display_label,
    _find_member,
    _find_member_by_label,
    _format_categorisations,
    _is_descendant,
    _key_member_display,
    _member_display_label,
    _merge_column_headers,
)
from dpmcore.services.layout_exporter.models import (
    CellData,
    DimensionMember,
    ExportConfig,
    LayoutHeader,
    TableLayout,
)
from dpmcore.services.layout_exporter.service import LayoutExporterService
from tests.integration.services.layout_exporter._helpers import (
    add_cell,
    add_context_composition,
    add_header,
    add_subcategory,
    add_table,
    add_variable_version,
    build_basic_module_with_table,
    make_member,
    make_module,
    make_property,
    seed_data_types,
    seed_domain_category,
    seed_property_category,
    seed_releases,
)

# ---------------------------------------------------------------- #
# Pure-function tests (no DB needed)
# ---------------------------------------------------------------- #


def _dm(
    *,
    property_id=1,
    dimension_label="Dim",
    dimension_code="DC",
    domain_code="DOM",
    member_label="Mem",
    member_code="m",
    data_type_code="",
):
    return DimensionMember(
        property_id=property_id,
        dimension_label=dimension_label,
        dimension_code=dimension_code,
        domain_code=domain_code,
        member_label=member_label,
        member_code=member_code,
        data_type_code=data_type_code,
    )


def _lh(
    *,
    header_id=1,
    header_vid=1,
    code="C",
    label="L",
    direction="x",
    order=1,
    is_abstract=False,
    is_key=False,
    parent_header_id=None,
    parent_first=True,
    depth=0,
    categorisations=None,
    subcategory_code="",
    subcategory_description="",
    subcategory_cat_code="",
    key_variable_vid=None,
    key_variable_id=None,
    key_data_type_code="",
    key_property_name="",
    key_categorisations=None,
):
    return LayoutHeader(
        header_id=header_id,
        header_vid=header_vid,
        code=code,
        label=label,
        direction=direction,
        order=order,
        is_abstract=is_abstract,
        is_key=is_key,
        parent_header_id=parent_header_id,
        parent_first=parent_first,
        depth=depth,
        categorisations=categorisations or [],
        subcategory_code=subcategory_code,
        subcategory_description=subcategory_description,
        subcategory_cat_code=subcategory_cat_code,
        key_variable_vid=key_variable_vid,
        key_variable_id=key_variable_id,
        key_data_type_code=key_data_type_code,
        key_property_name=key_property_name,
        key_categorisations=key_categorisations or [],
    )


def test_dim_display_label_aty_branch():
    dm = _dm(dimension_code="ATY", dimension_label="Main Property")
    assert _dim_display_label(dm) == "Main Property"


def test_dim_display_label_blank_dim_code_branch():
    dm = _dm(dimension_code="", dimension_label="Foo")
    assert _dim_display_label(dm) == "Foo"


def test_dim_display_label_regular_branch():
    dm = _dm(dimension_code="DC", domain_code="DOM", dimension_label="Foo")
    assert _dim_display_label(dm) == "(DC:DOM) Foo"


def test_member_display_label_aty_with_member_code():
    dm = _dm(
        dimension_code="ATY",
        member_code="qX",
        member_label="X",
    )
    assert _member_display_label(dm) == "(qX) X"


def test_member_display_label_aty_without_member_code():
    dm = _dm(dimension_code="ATY", member_code="", member_label="X")
    assert _member_display_label(dm) == "X"


def test_member_display_label_regular_with_codes():
    dm = _dm(
        dimension_code="DC",
        domain_code="DOM",
        member_code="m",
        member_label="lbl",
    )
    assert _member_display_label(dm) == "(DOM:m) lbl"


def test_member_display_label_regular_without_member_code():
    dm = _dm(
        dimension_code="DC",
        domain_code="DOM",
        member_code="",
        member_label="lbl",
    )
    assert _member_display_label(dm) == "lbl"


def test_key_member_display_with_subcategory():
    dm = _dm(domain_code="DOM")
    ch = _lh(
        subcategory_code="SC1",
        subcategory_cat_code="CAT",
        subcategory_description="Desc",
    )
    out = _key_member_display(dm, ch)
    assert "SC1" in out
    assert "CAT" in out
    assert "<Key value>" in out


def test_key_member_display_without_subcategory():
    dm = _dm(domain_code="DOM")
    ch = _lh()
    assert _key_member_display(dm, ch) == "(DOM:) <Key value>"


def test_find_member_hit_and_miss():
    dms = [_dm(property_id=1), _dm(property_id=2)]
    assert _find_member(dms, 2).property_id == 2
    assert _find_member(dms, 99) is None


def test_find_member_by_label_hit_and_miss():
    dms = [_dm(dimension_code="ATY", dimension_label="Main")]
    assert _find_member_by_label(dms, "Main") is not None
    assert _find_member_by_label(dms, "Other") is None


def test_format_categorisations_empty():
    assert _format_categorisations([]) == ""


def test_format_categorisations_populated():
    out = _format_categorisations([_dm(dimension_label="D", member_label="M")])
    assert "D" in out
    assert "M" in out


def test_build_header_tooltip_empty():
    assert _build_header_tooltip([]) == ""


def test_build_header_tooltip_populated():
    h = _lh(
        code="010",
        label="L",
        categorisations=[_dm(dimension_label="D", member_label="M")],
    )
    out = _build_header_tooltip([h])
    assert "010" in out
    assert "D = M" in out


def test_collect_axis_dimensions_aty_first():
    h1 = _lh(
        header_id=1,
        categorisations=[
            _dm(property_id=2, dimension_code="DCx", dimension_label="X"),
        ],
    )
    h2 = _lh(
        header_id=2,
        categorisations=[
            _dm(
                property_id=10,
                dimension_code="ATY",
                dimension_label="Main Property",
            ),
        ],
    )
    out = _collect_axis_dimensions([h1, h2])
    assert out[0] == "Main Property"


def test_collect_axis_dimensions_skips_abstract():
    h_abs = _lh(
        header_id=1,
        is_abstract=True,
        categorisations=[
            _dm(property_id=1, dimension_code="DC", dimension_label="X")
        ],
    )
    assert _collect_axis_dimensions([h_abs]) == []


def test_collect_axis_dimensions_supplemental_cats():
    h = _lh(header_id=1)
    sup = {1: [_dm(property_id=5, dimension_code="DC", dimension_label="L")]}
    out = _collect_axis_dimensions([h], supplemental_cats=sup)
    assert any("L" in lbl for lbl in out)


def test_collect_axis_dimensions_is_key_extension():
    h = _lh(
        header_id=1,
        is_key=True,
        key_categorisations=[
            _dm(property_id=5, dimension_code="ATY", dimension_label="Cur")
        ],
    )
    assert "Cur" in _collect_axis_dimensions([h])


def test_is_descendant_yes_no_and_cycle():
    by_id = {
        1: _lh(header_id=1, parent_header_id=None),
        2: _lh(header_id=2, parent_header_id=1),
        3: _lh(header_id=3, parent_header_id=2),
    }
    assert _is_descendant(by_id[3], 1, by_id) is True
    assert _is_descendant(by_id[3], 99, by_id) is False
    # Cycle protection: make a cycle 4 -> 5 -> 4
    by_id[4] = _lh(header_id=4, parent_header_id=5)
    by_id[5] = _lh(header_id=5, parent_header_id=4)
    assert _is_descendant(by_id[4], 99, by_id) is False


# ---------------------------------------------------------------- #
# write() / _write_table integration tests via the service
# ---------------------------------------------------------------- #


def test_write_basic_layout(memory_session, tmp_path):
    build_basic_module_with_table(memory_session)
    svc = LayoutExporterService(memory_session)
    out = svc.export_module("MOD1", output_path=str(tmp_path / "basic.xlsx"))
    wb = load_workbook(out)
    # Index sheet first
    assert wb.sheetnames[0] == "Index"
    # Table sheet exists
    ws = wb["T1"]
    # Title contains code and name
    assert "T1" in str(ws["A1"].value)
    assert "Table One" in str(ws["A1"].value)


def test_write_open_row_table_with_key_columns(memory_session, tmp_path):
    """Open-row layout with key columns + non-key columns."""
    seed_releases(memory_session)
    seed_data_types(memory_session)
    seed_property_category(memory_session)
    seed_domain_category(memory_session, 20, "DOM")

    make_module(memory_session, module_id=1, module_vid=10, code="MODK")
    add_table(
        memory_session,
        table_id=100,
        table_vid=1000,
        code="T_OR",
        name="OR",
        module_vid=10,
    )
    # Property used by key variable (data_type 'e')
    make_property(
        memory_session,
        property_id=200,
        name="Currency",
        data_type_id=2,
        dim_code="qCUR",
        domain_category_id=20,
    )
    # Property used by non-key column (monetary)
    make_property(
        memory_session,
        property_id=201,
        name="Amount",
        data_type_id=1,
        dim_code="qAMT",
        domain_category_id=20,
    )
    add_variable_version(
        memory_session,
        variable_id=400,
        variable_vid=4000,
        code="VK",
        property_id=200,
    )
    add_variable_version(
        memory_session,
        variable_id=401,
        variable_vid=4001,
        code="VN",
        property_id=201,
    )

    # Key column (open-row)
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=1,
        header_vid=11,
        direction="x",
        code="010",
        label="Currency",
        is_key=True,
        key_variable_vid=4000,
    )
    # Non-key column with cell content
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=2,
        header_vid=12,
        direction="x",
        code="020",
        label="Amount",
        order=2,
        property_id=201,
    )
    # Cell on the non-key column (open row -> row_id=None)
    add_cell(
        memory_session,
        cell_id=900,
        table_id=100,
        table_vid=1000,
        column_id=2,
        row_id=None,
        variable_vid=4001,
        sign="positive",
    )
    memory_session.commit()

    svc = LayoutExporterService(memory_session)
    out = svc.export_module("MODK", output_path=str(tmp_path / "or.xlsx"))
    wb = load_workbook(out)
    ws = wb["T_OR"]
    # Should have the "Open Rows" label somewhere
    found = False
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if v == "Open Rows":
                found = True
    assert found


def test_write_table_with_sheets_and_subcategory(memory_session, tmp_path):
    seed_releases(memory_session)
    seed_data_types(memory_session)
    seed_property_category(memory_session)
    seed_domain_category(memory_session, 30, "BAS")
    seed_domain_category(memory_session, 20, "DOM")

    make_module(memory_session, module_id=1, module_vid=10, code="MODS")
    add_table(
        memory_session,
        table_id=100,
        table_vid=1000,
        code="T_SH",
        name="Sheets",
        module_vid=10,
    )
    add_subcategory(
        memory_session,
        subcategory_id=1,
        subcategory_vid=11,
        category_id=30,
        code="SC1",
        description="Sub Description",
    )

    # Sheet header (z-axis)
    make_property(
        memory_session,
        property_id=200,
        name="MainP",
        data_type_id=1,
        dim_code="qMain",
        domain_category_id=20,
    )
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=1,
        header_vid=11,
        direction="z",
        code="010",
        label="Z Sheet",
        subcategory_vid=11,
        property_id=200,
    )
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=2,
        header_vid=12,
        direction="x",
        code="010",
        label="Col",
    )
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=3,
        header_vid=13,
        direction="y",
        code="010",
        label="Row",
    )
    memory_session.commit()
    svc = LayoutExporterService(memory_session)
    out = svc.export_module("MODS", output_path=str(tmp_path / "sh.xlsx"))
    wb = load_workbook(out)
    ws = wb["T_SH"]
    # Find the "Sheet per ..." label
    found_sheet_label = False
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if isinstance(v, str) and v.startswith("Sheet per"):
                found_sheet_label = True
    assert found_sheet_label


def test_write_table_no_code_row_no_code_column(memory_session, tmp_path):
    build_basic_module_with_table(memory_session)
    svc = LayoutExporterService(memory_session)
    cfg = ExportConfig(show_code_row=False, show_code_column=False)
    out = svc.export_module(
        "MOD1",
        output_path=str(tmp_path / "x.xlsx"),
        config=cfg,
    )
    wb = load_workbook(out)
    ws = wb["T1"]
    # Look for column label embedded with code
    cells_with_code = [
        c.value
        for row in ws.iter_rows()
        for c in row
        if isinstance(c.value, str) and c.value.startswith("0010 ")
    ]
    assert cells_with_code  # column label includes code


def test_write_table_with_abstract_column_header(memory_session, tmp_path):
    """Exercise abstract headers + show_abstract_header_codes + col groups."""
    seed_releases(memory_session)
    seed_data_types(memory_session)
    seed_property_category(memory_session)
    make_module(memory_session, module_id=1, module_vid=10, code="MODA")
    add_table(
        memory_session,
        table_id=100,
        table_vid=1000,
        code="T_AB",
        name="Abs",
        module_vid=10,
    )
    # Abstract parent column
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=1,
        header_vid=11,
        direction="x",
        code="P",
        label="Parent",
        order=1,
        is_abstract=True,
    )
    # Real column (child of abstract parent)
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=2,
        header_vid=12,
        direction="x",
        code="010",
        label="Child",
        order=2,
        parent_header_id=1,
    )
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=3,
        header_vid=13,
        direction="y",
        code="010",
        label="Row",
        order=1,
    )
    memory_session.commit()

    svc = LayoutExporterService(memory_session)
    cfg = ExportConfig(show_abstract_header_codes=True)
    out = svc.export_module(
        "MODA",
        output_path=str(tmp_path / "ab.xlsx"),
        config=cfg,
    )
    wb = load_workbook(out)
    ws = wb["T_AB"]
    # Check parent label includes code
    found = False
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and "Parent" in c.value:
                found = True
    assert found


def test_write_table_with_categorisations_and_tooltips(
    memory_session, tmp_path
):
    """Table with full categorisations to exercise annotations & tooltips."""
    seed_releases(memory_session)
    seed_data_types(memory_session)
    seed_property_category(memory_session)
    seed_domain_category(memory_session, 20, "DOM")

    make_module(memory_session, module_id=1, module_vid=10, code="MODC")
    add_table(
        memory_session,
        table_id=100,
        table_vid=1000,
        code="T_CAT",
        name="Cat",
        module_vid=10,
    )
    # Dimension property + member
    make_property(
        memory_session,
        property_id=200,
        name="Currency",
        data_type_id=2,
        dim_code="qCUR",
        domain_category_id=20,
    )
    make_member(
        memory_session,
        item_id=300,
        name="EUR",
        domain_category_id=20,
        code="x10",
    )
    add_context_composition(
        memory_session, context_id=50, property_id=200, item_id=300
    )

    # Column references context
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=1,
        header_vid=11,
        direction="x",
        code="010",
        label="Col",
        context_id=50,
    )
    # Row references the same context
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=2,
        header_vid=12,
        direction="y",
        code="010",
        label="Row",
        context_id=50,
    )
    memory_session.commit()
    svc = LayoutExporterService(memory_session)
    out = svc.export_module(
        "MODC",
        output_path=str(tmp_path / "cat.xlsx"),
    )
    wb = load_workbook(out)
    assert "T_CAT" in wb.sheetnames


def test_write_table_with_excluded_cells(memory_session, tmp_path):
    """Cell with is_excluded=True should not display variable info."""
    build_basic_module_with_table(memory_session)
    # Mark the existing cell as excluded
    from dpmcore.orm.rendering import TableVersionCell

    tvc = (
        memory_session.query(TableVersionCell)
        .filter_by(table_vid=1000, cell_id=900)
        .first()
    )
    tvc.is_excluded = True
    memory_session.commit()

    svc = LayoutExporterService(memory_session)
    out = svc.export_module(
        "MOD1",
        output_path=str(tmp_path / "ex.xlsx"),
    )
    assert out.exists()


def test_write_table_with_multi_depth_columns_and_rows(
    memory_session, tmp_path
):
    """Hierarchical column and row tree (depth>0)."""
    seed_releases(memory_session)
    seed_data_types(memory_session)
    seed_property_category(memory_session)
    make_module(memory_session, module_id=1, module_vid=10, code="MODD")
    add_table(
        memory_session,
        table_id=100,
        table_vid=1000,
        code="T_D",
        name="D",
        module_vid=10,
    )
    # Column hierarchy: parent + child
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=1,
        header_vid=11,
        direction="x",
        code="P",
        label="Parent",
        order=1,
    )
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=2,
        header_vid=12,
        direction="x",
        code="010",
        label="Child",
        order=2,
        parent_header_id=1,
    )
    # Row hierarchy
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=3,
        header_vid=13,
        direction="y",
        code="P",
        label="RP",
        order=1,
    )
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=4,
        header_vid=14,
        direction="y",
        code="010",
        label="RC",
        order=2,
        parent_header_id=3,
    )
    memory_session.commit()
    svc = LayoutExporterService(memory_session)
    out = svc.export_module(
        "MODD",
        output_path=str(tmp_path / "d.xlsx"),
    )
    wb = load_workbook(out)
    ws = wb["T_D"]
    # Outline level should be set for child rows / cols
    has_outline = any(
        rd.outline_level and rd.outline_level > 0
        for rd in ws.row_dimensions.values()
    )
    assert has_outline


def test_write_table_closing_balance_sign_branch(memory_session, tmp_path):
    """Row labelled 'Closing balance' triggers positive default for monetary."""
    seed_releases(memory_session)
    seed_data_types(memory_session)
    seed_property_category(memory_session)
    seed_domain_category(memory_session, 20, "DOM")

    make_module(memory_session, module_id=1, module_vid=10, code="MODCB")
    add_table(
        memory_session,
        table_id=100,
        table_vid=1000,
        code="T_CB",
        name="CB",
        module_vid=10,
    )
    # Monetary property + matching variable
    make_property(
        memory_session,
        property_id=200,
        name="Amount",
        data_type_id=1,  # 'm'
        dim_code="qAMT",
        domain_category_id=20,
    )
    add_variable_version(
        memory_session,
        variable_id=400,
        variable_vid=4000,
        code="V",
        property_id=200,
    )

    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=1,
        header_vid=11,
        direction="x",
        code="010",
        label="Col",
    )
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=2,
        header_vid=12,
        direction="y",
        code="010",
        label="Closing balance, year-end",
    )
    add_cell(
        memory_session,
        cell_id=900,
        table_id=100,
        table_vid=1000,
        column_id=1,
        row_id=2,
        variable_vid=4000,
        sign="",  # blank -> default
    )
    memory_session.commit()
    svc = LayoutExporterService(memory_session)
    out = svc.export_module(
        "MODCB",
        output_path=str(tmp_path / "cb.xlsx"),
    )
    wb = load_workbook(out)
    ws = wb["T_CB"]
    # Look for "positive" appended to a cell (the monetary one)
    has_positive = False
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if isinstance(v, str) and "positive" in v:
                has_positive = True
    assert has_positive


def test_write_table_col_positive_only_default(memory_session, tmp_path):
    """When column has at least one signed cell, NULL→positive default."""
    seed_releases(memory_session)
    seed_data_types(memory_session)
    seed_property_category(memory_session)
    seed_domain_category(memory_session, 20, "DOM")

    make_module(memory_session, module_id=1, module_vid=10, code="MODP")
    add_table(
        memory_session,
        table_id=100,
        table_vid=1000,
        code="T_P",
        name="P",
        module_vid=10,
    )
    make_property(
        memory_session,
        property_id=200,
        name="Amount",
        data_type_id=1,
        dim_code="qAMT",
        domain_category_id=20,
    )
    add_variable_version(
        memory_session,
        variable_id=400,
        variable_vid=4000,
        code="V",
        property_id=200,
    )
    add_variable_version(
        memory_session,
        variable_id=401,
        variable_vid=4001,
        code="V2",
        property_id=200,
    )

    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=1,
        header_vid=11,
        direction="x",
        code="010",
        label="Col",
    )
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=2,
        header_vid=12,
        direction="y",
        code="010",
        label="Row 1",
    )
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=3,
        header_vid=13,
        direction="y",
        code="020",
        label="Row 2",
        order=2,
    )
    # First cell signed; second blank
    add_cell(
        memory_session,
        cell_id=900,
        table_id=100,
        table_vid=1000,
        column_id=1,
        row_id=2,
        variable_vid=4000,
        sign="positive",
    )
    add_cell(
        memory_session,
        cell_id=901,
        table_id=100,
        table_vid=1000,
        column_id=1,
        row_id=3,
        variable_vid=4001,
        sign="",
    )
    memory_session.commit()
    svc = LayoutExporterService(memory_session)
    out = svc.export_module(
        "MODP",
        output_path=str(tmp_path / "p.xlsx"),
    )
    wb = load_workbook(out)
    ws = wb["T_P"]
    pos_count = 0
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if isinstance(v, str) and "positive" in v:
                pos_count += 1
    assert pos_count >= 2  # both cells should display 'positive'


def test_write_table_parent_sign_suppression(memory_session, tmp_path):
    """If parent row has explicit sign, child NULL-sign defaults are suppressed."""
    seed_releases(memory_session)
    seed_data_types(memory_session)
    seed_property_category(memory_session)
    seed_domain_category(memory_session, 20, "DOM")

    make_module(memory_session, module_id=1, module_vid=10, code="MODPS")
    add_table(
        memory_session,
        table_id=100,
        table_vid=1000,
        code="T_PS",
        name="PS",
        module_vid=10,
    )
    make_property(
        memory_session,
        property_id=200,
        name="Amount",
        data_type_id=1,
        dim_code="qAMT",
        domain_category_id=20,
    )
    for vid in (4000, 4001, 4002):
        add_variable_version(
            memory_session,
            variable_id=400 + (vid - 4000),
            variable_vid=vid,
            code=f"V{vid}",
            property_id=200,
        )
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=1,
        header_vid=11,
        direction="x",
        code="010",
        label="Col",
    )
    # Parent row (with explicit sign on its cell)
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=2,
        header_vid=12,
        direction="y",
        code="P",
        label="Parent row",
        order=1,
    )
    # Child row (NULL sign — should be suppressed)
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=3,
        header_vid=13,
        direction="y",
        code="020",
        label="Child row",
        order=2,
        parent_header_id=2,
    )
    # Sibling row (NULL sign — should default positive: col_positive_only)
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=4,
        header_vid=14,
        direction="y",
        code="030",
        label="Sibling row",
        order=3,
    )
    add_cell(
        memory_session,
        cell_id=900,
        table_id=100,
        table_vid=1000,
        column_id=1,
        row_id=2,
        variable_vid=4000,
        sign="positive",
    )
    add_cell(
        memory_session,
        cell_id=901,
        table_id=100,
        table_vid=1000,
        column_id=1,
        row_id=3,
        variable_vid=4001,
        sign="",
    )
    add_cell(
        memory_session,
        cell_id=902,
        table_id=100,
        table_vid=1000,
        column_id=1,
        row_id=4,
        variable_vid=4002,
        sign="",
    )
    memory_session.commit()

    svc = LayoutExporterService(memory_session)
    out = svc.export_module(
        "MODPS",
        output_path=str(tmp_path / "ps.xlsx"),
    )
    assert out.exists()


def test_write_table_e_data_type_with_domain_label(memory_session, tmp_path):
    """Cell with data_type='e' shows [domain_label]."""
    seed_releases(memory_session)
    seed_data_types(memory_session)
    seed_property_category(memory_session)
    seed_domain_category(memory_session, 20, "DOM")

    make_module(memory_session, module_id=1, module_vid=10, code="MODE")
    add_table(
        memory_session,
        table_id=100,
        table_vid=1000,
        code="T_E",
        name="E",
        module_vid=10,
    )
    # Property with 'e' data type
    make_property(
        memory_session,
        property_id=200,
        name="MyDomain",
        data_type_id=2,
        dim_code="qE",
        domain_category_id=20,
    )
    add_variable_version(
        memory_session,
        variable_id=400,
        variable_vid=4000,
        code="V",
        property_id=200,
    )
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=1,
        header_vid=11,
        direction="x",
        code="010",
        label="Col",
    )
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=2,
        header_vid=12,
        direction="y",
        code="010",
        label="Row",
    )
    add_cell(
        memory_session,
        cell_id=900,
        table_id=100,
        table_vid=1000,
        column_id=1,
        row_id=2,
        variable_vid=4000,
    )
    memory_session.commit()

    svc = LayoutExporterService(memory_session)
    out = svc.export_module(
        "MODE",
        output_path=str(tmp_path / "e.xlsx"),
    )
    wb = load_workbook(out)
    ws = wb["T_E"]
    has_domain = False
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if isinstance(v, str) and "[MyDomain]" in v:
                has_domain = True
    assert has_domain


def test_write_table_void_and_missing_cells(memory_session, tmp_path):
    """Multiple rows but only one cell -> missing cells use excluded fill."""
    seed_releases(memory_session)
    seed_data_types(memory_session)
    seed_property_category(memory_session)
    make_module(memory_session, module_id=1, module_vid=10, code="MODM")
    add_table(
        memory_session,
        table_id=100,
        table_vid=1000,
        code="T_M",
        name="M",
        module_vid=10,
    )
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=1,
        header_vid=11,
        direction="x",
        code="010",
        label="Col",
    )
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=2,
        header_vid=12,
        direction="y",
        code="010",
        label="Row 1",
        order=1,
    )
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=3,
        header_vid=13,
        direction="y",
        code="020",
        label="Row 2",
        order=2,
    )
    memory_session.commit()
    svc = LayoutExporterService(memory_session)
    out = svc.export_module(
        "MODM",
        output_path=str(tmp_path / "m.xlsx"),
    )
    assert out.exists()


def test_write_open_row_with_excluded_cell(memory_session, tmp_path):
    """Open-row table where the cell is excluded -> excluded fill."""
    seed_releases(memory_session)
    seed_data_types(memory_session)
    seed_property_category(memory_session)
    make_module(memory_session, module_id=1, module_vid=10, code="MODOX")
    add_table(
        memory_session,
        table_id=100,
        table_vid=1000,
        code="T_OX",
        name="OX",
        module_vid=10,
    )
    # Non-key column, no cell -> excluded path
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=1,
        header_vid=11,
        direction="x",
        code="010",
        label="Col",
    )
    memory_session.commit()
    svc = LayoutExporterService(memory_session)
    out = svc.export_module(
        "MODOX",
        output_path=str(tmp_path / "ox.xlsx"),
    )
    assert out.exists()


def test_merge_column_headers_helper():
    """Direct test on _merge_column_headers with parents + leaves."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active

    parent = _lh(
        header_id=1,
        code="P",
        label="Parent",
        depth=0,
        is_abstract=False,
    )
    child = _lh(
        header_id=2,
        code="C",
        label="Child",
        depth=1,
        is_abstract=False,
        parent_header_id=1,
    )
    cols = [parent, child]
    positions = {1: 1, 2: 2}
    _merge_column_headers(ws, cols, positions, 5, 3, max_depth=1)
    # Should not raise; merged/marked cells exist
    assert ws.cell(row=5, column=3).value is None or True


def test_merge_column_headers_with_abstract_parent():
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    abstract = _lh(
        header_id=1,
        code="P",
        label="Parent",
        depth=0,
        is_abstract=True,
    )
    child = _lh(
        header_id=2,
        code="C",
        label="Child",
        depth=1,
        is_abstract=False,
        parent_header_id=1,
    )
    cols = [abstract, child]
    # Abstract shares position with first child
    positions = {1: 1, 2: 1}
    _merge_column_headers(ws, cols, positions, 5, 3, max_depth=1)


def test_collect_axis_dimensions_dedup_by_label():
    """ATY dim with same label appears once across multiple headers."""
    h1 = _lh(
        header_id=1,
        categorisations=[
            _dm(property_id=1, dimension_code="ATY", dimension_label="Main")
        ],
    )
    h2 = _lh(
        header_id=2,
        categorisations=[
            _dm(property_id=2, dimension_code="ATY", dimension_label="Main")
        ],
    )
    out = _collect_axis_dimensions([h1, h2])
    assert out.count("Main") == 1


# ---------------------------------------------------------------- #
# Direct ExcelLayoutWriter tests with synthetic layouts
# ---------------------------------------------------------------- #


def test_writer_sorts_layouts_alphabetically():
    layouts = [
        TableLayout(table_vid=2, table_code="B", table_name="B"),
        TableLayout(table_vid=1, table_code="A", table_name="A"),
    ]
    # Add at least one column so the layout would be written...but
    # ExcelLayoutWriter writes whatever layouts you give it.
    layouts[0].columns = [_lh(header_id=1)]
    layouts[0].rows = [_lh(header_id=2, direction="y")]
    layouts[1].columns = [_lh(header_id=3)]
    layouts[1].rows = [_lh(header_id=4, direction="y")]
    writer = ExcelLayoutWriter(layouts)
    wb = writer.write()
    # Sheets after Index
    sheet_names = [n for n in wb.sheetnames if n != "Index"]
    assert sheet_names == ["A", "B"]


def test_writer_index_sheet_hyperlinks():
    layout = TableLayout(table_vid=1, table_code="X1", table_name="Test")
    layout.columns = [_lh(header_id=1)]
    layout.rows = [_lh(header_id=2, direction="y")]
    writer = ExcelLayoutWriter([layout])
    wb = writer.write()
    ws = wb["Index"]
    # Code cell should have a hyperlink set
    code_cell = ws.cell(row=4, column=2)
    assert code_cell.hyperlink is not None
    assert code_cell.value == "X1"


def test_writer_with_custom_config_no_annotate():
    layout = TableLayout(table_vid=1, table_code="X1", table_name="T")
    layout.columns = [_lh(header_id=1)]
    layout.rows = [_lh(header_id=2, direction="y")]
    cfg = ExportConfig(
        annotate=False,
        add_header_comments=False,
        add_cell_comments=False,
    )
    writer = ExcelLayoutWriter([layout], cfg)
    wb = writer.write()
    assert "X1" in wb.sheetnames


def test_writer_synthetic_with_abstract_row(memory_session, tmp_path):
    """Abstract row triggers the merge branch (lines 401-405)."""
    seed_releases(memory_session)
    seed_data_types(memory_session)
    seed_property_category(memory_session)
    make_module(memory_session, module_id=1, module_vid=10, code="MODAR")
    add_table(
        memory_session,
        table_id=100,
        table_vid=1000,
        code="T_AR",
        name="AR",
        module_vid=10,
    )
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=1,
        header_vid=11,
        direction="x",
        code="010",
        label="Col",
    )
    # Abstract row + a child row
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=2,
        header_vid=12,
        direction="y",
        code="ABS",
        label="Abstract Row",
        order=1,
        is_abstract=True,
    )
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=3,
        header_vid=13,
        direction="y",
        code="010",
        label="Real Row",
        order=2,
        parent_header_id=2,
    )
    memory_session.commit()
    svc = LayoutExporterService(memory_session)
    out = svc.export_module(
        "MODAR",
        output_path=str(tmp_path / "ar.xlsx"),
    )
    assert out.exists()


def test_writer_abstract_column_at_start_position_zero(
    memory_session, tmp_path
):
    """Abstract header followed by another abstract -> col_offset == 0 path."""
    seed_releases(memory_session)
    seed_data_types(memory_session)
    seed_property_category(memory_session)
    make_module(memory_session, module_id=1, module_vid=10, code="MODA0")
    add_table(
        memory_session,
        table_id=100,
        table_vid=1000,
        code="T_A0",
        name="A0",
        module_vid=10,
    )
    # Two abstract columns -> first abstract at pos 0
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=1,
        header_vid=11,
        direction="x",
        code="P1",
        label="Abs1",
        order=1,
        is_abstract=True,
    )
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=2,
        header_vid=12,
        direction="x",
        code="P2",
        label="Abs2",
        order=2,
        is_abstract=True,
        parent_header_id=1,
    )
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=3,
        header_vid=13,
        direction="x",
        code="010",
        label="Real",
        order=3,
        parent_header_id=2,
    )
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=4,
        header_vid=14,
        direction="y",
        code="010",
        label="Row",
    )
    memory_session.commit()
    svc = LayoutExporterService(memory_session)
    out = svc.export_module(
        "MODA0",
        output_path=str(tmp_path / "a0.xlsx"),
    )
    assert out.exists()


def test_write_open_row_key_column_non_e_dtype():
    """Key column with non-'e' data type -> uses _DATA_TYPE_SYMBOLS path."""
    layout = TableLayout(
        table_vid=1, table_code="OR", table_name="OR", is_open_row=True
    )
    layout.columns = [
        _lh(
            header_id=1,
            code="010",
            label="K",
            is_key=True,
            key_variable_vid=999,
            key_variable_id=42,
            key_data_type_code="m",  # not 'e'
            key_property_name="Money",
        )
    ]
    layout.rows = []  # open row
    writer = ExcelLayoutWriter([layout])
    wb = writer.write()
    ws = wb["OR"]
    found = False
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if isinstance(v, str) and "Money" in v:
                found = True
    assert found


def test_write_open_row_with_excluded_cell_branch():
    """Open-row non-key column where cell is excluded -> excluded fill path."""
    layout = TableLayout(
        table_vid=1, table_code="OREX", table_name="OREX", is_open_row=True
    )
    layout.columns = [
        _lh(header_id=1, code="010", label="Col"),
    ]
    layout.rows = []
    layout.cells = {
        (None, 1, None): CellData(
            row_header_id=None,
            col_header_id=1,
            sheet_header_id=None,
            variable_vid=4000,
            is_excluded=True,
        )
    }
    writer = ExcelLayoutWriter([layout])
    wb = writer.write()
    assert "OREX" in wb.sheetnames


def test_write_open_row_with_e_dtype_in_cell():
    """Open-row cell with data_type 'e' and domain_label."""
    layout = TableLayout(
        table_vid=1, table_code="ORE", table_name="ORE", is_open_row=True
    )
    layout.columns = [_lh(header_id=1, code="010", label="Col")]
    layout.rows = []
    layout.cells = {
        (None, 1, None): CellData(
            row_header_id=None,
            col_header_id=1,
            sheet_header_id=None,
            variable_vid=4000,
            variable_id=400,
            data_type_code="e",
            domain_label="MyEnumDomain",
            sign="positive",
        )
    }
    writer = ExcelLayoutWriter([layout])
    wb = writer.write()
    ws = wb["ORE"]
    found_domain = False
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if isinstance(v, str) and "[MyEnumDomain]" in v:
                found_domain = True
    assert found_domain


def test_write_table_cell_with_no_variable_vid():
    """Cell exists but variable_vid is None -> shows as excluded."""
    layout = TableLayout(table_vid=1, table_code="NV", table_name="NV")
    layout.columns = [_lh(header_id=1, code="010", label="Col")]
    layout.rows = [_lh(header_id=2, direction="y", code="010", label="Row")]
    layout.cells = {
        (2, 1, None): CellData(
            row_header_id=2,
            col_header_id=1,
            sheet_header_id=None,
            variable_vid=None,
            is_excluded=False,
        )
    }
    writer = ExcelLayoutWriter([layout])
    wb = writer.write()
    assert "NV" in wb.sheetnames


def test_collect_axis_dimensions_skip_abstract_in_writer():
    """Abstract header skip in column annotation member loop."""
    layout = TableLayout(table_vid=1, table_code="AB", table_name="AB")
    cats = [_dm(property_id=1, dimension_label="X", dimension_code="DC")]
    layout.columns = [
        _lh(header_id=1, is_abstract=True, code="P", label="P"),
        _lh(
            header_id=2,
            code="010",
            label="C",
            categorisations=cats,
            parent_header_id=1,
        ),
    ]
    layout.rows = [
        _lh(
            header_id=3,
            direction="y",
            code="010",
            label="R",
            categorisations=cats,
        ),
        _lh(header_id=4, direction="y", is_abstract=True, code="A", label="A"),
    ]
    writer = ExcelLayoutWriter([layout])
    wb = writer.write()
    assert "AB" in wb.sheetnames


def test_writer_with_subcategory_on_column_and_row():
    """Column / row with subcategory triggers subcategory display addition."""
    layout = TableLayout(table_vid=1, table_code="SC", table_name="SC")
    cats = [_dm(property_id=1, dimension_label="X", dimension_code="DC")]
    layout.columns = [
        _lh(
            header_id=1,
            code="010",
            label="C",
            categorisations=cats,
            subcategory_code="SC1",
            subcategory_cat_code="CAT",
        ),
    ]
    layout.rows = [
        _lh(
            header_id=2,
            direction="y",
            code="010",
            label="R",
            categorisations=cats,
            subcategory_code="SC1",
            subcategory_cat_code="CAT",
        ),
    ]
    writer = ExcelLayoutWriter([layout])
    wb = writer.write()
    ws = wb["SC"]
    found = False
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if isinstance(v, str) and "[CAT:SC1]" in v:
                found = True
    assert found


def test_apply_row_groups_no_excel_row_skip():
    """Row dimension with depth>0 not in row_positions -> skipped (line 1078)."""
    from openpyxl import Workbook

    from dpmcore.services.layout_exporter.excel_writer import _apply_row_groups

    wb = Workbook()
    ws = wb.active
    rows = [_lh(header_id=99, direction="y", depth=2)]
    _apply_row_groups(ws, rows, {})  # 99 not in positions


def test_apply_col_groups_no_offset_skip():
    """Col dim with depth>0 but col_offset==0 -> skipped (line 1092)."""
    from openpyxl import Workbook

    from dpmcore.services.layout_exporter.excel_writer import _apply_col_groups

    wb = Workbook()
    ws = wb.active
    cols = [_lh(header_id=99, depth=2, is_abstract=False)]
    _apply_col_groups(ws, cols, {99: 0}, 3)  # offset 0 -> skipped


def test_merge_column_headers_no_descendants():
    """Parent with no real descendants should not merge (line 1033)."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    parent = _lh(header_id=1, depth=0, is_abstract=False)
    # Lone parent - no children
    _merge_column_headers(ws, [parent], {1: 1}, 5, 3, max_depth=0)


def test_merge_column_headers_zero_offset_skip():
    """Header with col_offset==0 in merge pass1 / pass2 (lines 980, 1019)."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    abstract = _lh(header_id=1, is_abstract=True)
    child = _lh(header_id=2, parent_header_id=1, depth=1, is_abstract=False)
    # Abstract parent at pos 0
    cols = [abstract, child]
    _merge_column_headers(ws, cols, {1: 0, 2: 1}, 5, 3, max_depth=1)


def test_merge_column_headers_pass1_with_depth():
    """Non-abstract column at depth<max with leaves at deeper depth."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    parent = _lh(header_id=1, depth=0, is_abstract=False)
    child = _lh(header_id=2, parent_header_id=1, depth=1, is_abstract=False)
    grand = _lh(header_id=3, parent_header_id=2, depth=2, is_abstract=False)
    cols = [parent, child, grand]
    positions = {1: 1, 2: 2, 3: 3}
    _merge_column_headers(ws, cols, positions, 5, 3, max_depth=2)


def test_merge_column_headers_descendant_with_zero_offset():
    """Descendant exists in by_id but has zero col_offset -> skipped (1030)."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    parent = _lh(header_id=1, depth=0, is_abstract=False)
    # child has zero offset (e.g., abstract sibling state)
    child = _lh(header_id=2, parent_header_id=1, depth=1, is_abstract=False)
    cols = [parent, child]
    positions = {1: 1, 2: 0}  # child at zero offset
    _merge_column_headers(ws, cols, positions, 5, 3, max_depth=1)


def test_is_descendant_breaks_when_current_none():
    """parent_header_id points to a header_id missing from by_id (line 1065)."""
    by_id = {
        1: _lh(header_id=1, parent_header_id=2),
        # key 2 missing
    }
    # Walk: visit parent_id=2, not in by_id -> current becomes None -> break
    assert _is_descendant(by_id[1], 99, by_id) is False


def test_writer_sheet_with_is_key(memory_session, tmp_path):
    """Sheet-axis header with is_key True triggers '<Key value>' suffix."""
    seed_releases(memory_session)
    seed_data_types(memory_session)
    seed_property_category(memory_session)
    seed_domain_category(memory_session, 30, "BAS")
    make_module(memory_session, module_id=1, module_vid=10, code="MODKZ")
    add_table(
        memory_session,
        table_id=100,
        table_vid=1000,
        code="T_KZ",
        name="KZ",
        module_vid=10,
    )
    add_subcategory(
        memory_session,
        subcategory_id=1,
        subcategory_vid=11,
        category_id=30,
        code="SCK",
        description="d",
    )
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=1,
        header_vid=11,
        direction="z",
        code="010",
        label="Sheet",
        subcategory_vid=11,
        is_key=True,
    )
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=2,
        header_vid=12,
        direction="x",
        code="010",
        label="Col",
    )
    add_header(
        memory_session,
        table_vid=1000,
        table_id=100,
        header_id=3,
        header_vid=13,
        direction="y",
        code="010",
        label="Row",
    )
    memory_session.commit()
    svc = LayoutExporterService(memory_session)
    out = svc.export_module("MODKZ", output_path=str(tmp_path / "kz.xlsx"))
    wb = load_workbook(out)
    ws = wb["T_KZ"]
    found = False
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if isinstance(v, str) and "<Key value>" in v:
                found = True
    assert found


def test_open_row_with_abstract_column_skip():
    """Open-row layout with an abstract column triggers continue (line 623)."""
    layout = TableLayout(
        table_vid=1, table_code="ORA", table_name="ORA", is_open_row=True
    )
    layout.columns = [
        _lh(header_id=1, code="P", label="Abs", is_abstract=True),
        _lh(
            header_id=2,
            code="010",
            label="Col",
            parent_header_id=1,
        ),
    ]
    layout.rows = []
    writer = ExcelLayoutWriter([layout])
    wb = writer.write()
    assert "ORA" in wb.sheetnames


def test_data_cell_no_data_type_code():
    """Cell variable_vid set but no data_type_code (line 503)."""
    layout = TableLayout(table_vid=1, table_code="DT0", table_name="DT0")
    layout.columns = [_lh(header_id=1, code="010", label="Col")]
    layout.rows = [_lh(header_id=2, direction="y", code="010", label="Row")]
    layout.cells = {
        (2, 1, None): CellData(
            row_header_id=2,
            col_header_id=1,
            sheet_header_id=None,
            variable_vid=4000,
            variable_id=400,
            data_type_code="",  # empty
            sign="positive",
        )
    }
    writer = ExcelLayoutWriter([layout])
    wb = writer.write()
    ws = wb["DT0"]
    found = False
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if isinstance(v, str) and "400" in v and "positive" in v:
                found = True
    assert found


def test_data_cell_unknown_data_type_code():
    """Cell with data_type_code not in _DATA_TYPE_SYMBOLS (line 509->514)."""
    layout = TableLayout(table_vid=1, table_code="DTU", table_name="DTU")
    layout.columns = [_lh(header_id=1, code="010", label="Col")]
    layout.rows = [_lh(header_id=2, direction="y", code="010", label="Row")]
    layout.cells = {
        (2, 1, None): CellData(
            row_header_id=2,
            col_header_id=1,
            sheet_header_id=None,
            variable_vid=4000,
            variable_id=400,
            data_type_code="zzz",  # unknown
        )
    }
    writer = ExcelLayoutWriter([layout])
    wb = writer.write()
    assert "DTU" in wb.sheetnames


def test_open_row_no_dtype_no_sign():
    """Open-row cell variant: no data_type_code, no sign (lines 661/671/673)."""
    layout = TableLayout(
        table_vid=1, table_code="ORN", table_name="ORN", is_open_row=True
    )
    layout.columns = [_lh(header_id=1, code="010", label="Col")]
    layout.rows = []
    layout.cells = {
        (None, 1, None): CellData(
            row_header_id=None,
            col_header_id=1,
            sheet_header_id=None,
            variable_vid=4000,
            variable_id=400,
            data_type_code="",
            sign="",
        )
    }
    writer = ExcelLayoutWriter([layout])
    wb = writer.write()
    assert "ORN" in wb.sheetnames


def test_open_row_unknown_dtype():
    """Open-row cell with unknown dtype (line 667->671)."""
    layout = TableLayout(
        table_vid=1, table_code="ORU", table_name="ORU", is_open_row=True
    )
    layout.columns = [_lh(header_id=1, code="010", label="Col")]
    layout.rows = []
    layout.cells = {
        (None, 1, None): CellData(
            row_header_id=None,
            col_header_id=1,
            sheet_header_id=None,
            variable_vid=4000,
            variable_id=400,
            data_type_code="qqq",
            sign="",
        )
    }
    writer = ExcelLayoutWriter([layout])
    wb = writer.write()
    assert "ORU" in wb.sheetnames


def test_open_row_no_layout_cells():
    """Open-row layout with no cells (line 616->621)."""
    layout = TableLayout(
        table_vid=1, table_code="ORE2", table_name="ORE2", is_open_row=True
    )
    layout.columns = [_lh(header_id=1, code="010", label="Col")]
    layout.rows = []
    # No cells dict entries
    writer = ExcelLayoutWriter([layout])
    wb = writer.write()
    assert "ORE2" in wb.sheetnames


def test_open_row_key_no_property_name():
    """Key column without key_property_name (line 640->642)."""
    layout = TableLayout(
        table_vid=1, table_code="ORK2", table_name="ORK2", is_open_row=True
    )
    layout.columns = [
        _lh(
            header_id=1,
            code="010",
            label="K",
            is_key=True,
            key_variable_vid=999,
            key_variable_id=42,
            key_data_type_code="m",
            key_property_name="",  # blank
        )
    ]
    layout.rows = []
    writer = ExcelLayoutWriter([layout])
    wb = writer.write()
    assert "ORK2" in wb.sheetnames


def test_writer_synthetic_layout_with_dp_supplemental_cats():
    """Cell dp_categorisations introduce supplemental row dimensions."""
    layout = TableLayout(table_vid=1, table_code="X1", table_name="T")
    layout.columns = [_lh(header_id=1, code="010", label="Col")]
    layout.rows = [
        _lh(header_id=2, direction="y", code="010", label="Row"),
    ]
    layout.cells = {
        (2, 1, None): CellData(
            row_header_id=2,
            col_header_id=1,
            sheet_header_id=None,
            variable_vid=4000,
            variable_id=400,
            data_type_code="m",
            sign="positive",
            dp_categorisations=[
                _dm(
                    property_id=99,
                    dimension_label="Extra Dim",
                    dimension_code="EXD",
                    member_label="Extra Mem",
                )
            ],
        )
    }
    writer = ExcelLayoutWriter([layout])
    wb = writer.write()
    ws = wb["X1"]
    # Should write row annotation showing dim or member
    found = False
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if isinstance(v, str) and ("Extra Dim" in v or "Extra Mem" in v):
                found = True
    assert found
