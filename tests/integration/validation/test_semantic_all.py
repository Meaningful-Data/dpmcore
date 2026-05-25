"""Parametrized semantic tests for all expressions in the validations xlsx.

Skipped automatically when tests/fixtures/validations_export.xlsx or
tests/fixtures/test_data.db is absent.
"""

from pathlib import Path

import openpyxl
import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from dpmcore.services.semantic import SemanticService

_XLSX = (
    Path(__file__).resolve().parents[3]
    / "tests"
    / "fixtures"
    / "validations_export.xlsx"
)
_DB = (
    Path(__file__).resolve().parents[3] / "tests" / "fixtures" / "test_data.db"
)


def _missing_reason() -> str | None:
    if not _XLSX.exists():
        return f"xlsx not found: {_XLSX}"
    if not _DB.exists():
        return f"db not found: {_DB}"
    return None


_MISSING = _missing_reason()


def _load_params():
    if _MISSING:
        return [
            pytest.param(
                "",
                "",
                "",
                "",
                marks=pytest.mark.skip(reason=_MISSING),
            )
        ]
    wb = openpyxl.load_workbook(_XLSX)
    ws = wb["Validations"]
    headers = [cell.value for cell in ws[1]]
    params = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row, strict=False))
        code = str(d.get("Code") or "")
        release = str(d.get("StartRelease") or "").strip()
        for col in ("Expression", "Precondition"):
            val = d.get(col)
            if val and str(val).strip():
                params.append(
                    pytest.param(
                        code,
                        col,
                        release,
                        str(val).strip(),
                        id=f"{code}-{col}",
                    )
                )
    return params


@pytest.fixture(scope="module")
def semantic_service():
    if _MISSING:
        pytest.skip(_MISSING)
    engine = create_engine(f"sqlite:///{_DB}")
    Session = sessionmaker(bind=engine)
    session = Session()
    svc = SemanticService(session)
    yield svc
    session.close()
    engine.dispose()


@pytest.mark.parametrize(
    ("code", "col", "release", "expression"), _load_params()
)
def test_semantic(code, col, release, expression, semantic_service):
    result = semantic_service.validate(
        expression, release_code=release or None
    )
    assert result.is_valid, f"{code} | {col} | {result.error_message}"
