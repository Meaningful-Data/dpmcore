"""Excel workbook writer for table layouts.

Takes processed TableLayout dataclasses and renders them to an
openpyxl Workbook with formatting, merged headers, annotations,
comments, and outline groups.
"""

from __future__ import annotations

from typing import Any, NamedTuple, Optional

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

from dpmcore.services.layout_exporter.models import (
    CellData,
    DimensionMember,
    ExportConfig,
    LayoutHeader,
    TableLayout,
)

# --- Style constants ---

_THIN = Side(style="thin")
_BORDER_ALL = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_GREY_FILL = PatternFill(
    start_color="D8D8D8", end_color="D8D8D8", fill_type="solid"
)
_EXCLUDED_FILL = PatternFill(
    start_color="999999", end_color="999999", fill_type="solid"
)
# Void cells are excluded too, so they get a darker grey to stand out
# from the cells that are merely not reportable.
_VOID_FILL = PatternFill(
    start_color="595959", end_color="595959", fill_type="solid"
)
# Cells whose datapoint is shared with another cell of the workbook:
# "identities" in DPM terms — the same figure reported in two places.
_IDENTITY_FILL = PatternFill(
    start_color="FFFF00", end_color="FFFF00", fill_type="solid"
)
# Key cells of open tables: the variables that identify a reported row
# rather than carrying a reported figure.
_KEY_FILL = PatternFill(
    start_color="C4D79B", end_color="C4D79B", fill_type="solid"
)
# Excel worksheet titles are limited to 31 characters.
_MAX_SHEET_TITLE = 31
# Cap the identity list in a tooltip; a few datapoints are reported in
# dozens of cells and the comment box would be unreadable.
_MAX_IDENTITIES_IN_TOOLTIP = 15

_TITLE_FONT = Font(bold=True, size=11)
_HEADER_FONT = Font(bold=True, size=10)
_DATA_FONT = Font(size=9)

# Data type code -> display symbol (matching EBA convention)
_DATA_TYPE_SYMBOLS: dict[str, str] = {
    "b": "TRUE/FALSE",
    "t": "TRUE",
    "d": "yyyy-mm-dd",
    "m": "\u20ac\u00a3$",  # €£$
    "p": "%",
    "i": "#",
    "r": "#.###",
    "s": "text",
}

# Dimension annotation colors (cycling palette)
_DIM_COLORS = [
    "0000CC",
    "CC0000",
    "006600",
    "990099",
    "CC6600",
    "006666",
    "660066",
    "336699",
    "993333",
    "339933",
    "666699",
    "996633",
    "339966",
    "993366",
    "669933",
]


class _IndexEntry(NamedTuple):
    """One worksheet written for a table (one per Z-axis sheet)."""

    title: str
    layout: TableLayout
    sheet_label: str


class ExcelLayoutWriter:
    """Writes one or more TableLayouts to an openpyxl Workbook."""

    def __init__(
        self,
        layouts: list[TableLayout],
        config: Optional[ExportConfig] = None,
    ) -> None:
        """Initialise with the layouts to render and optional config."""
        self.layouts = layouts
        self.config = config or ExportConfig()
        # {variable_vid: [cell location, ...]} for shared datapoints;
        # filled in write().
        self.identities: dict[int, list[str]] = {}
        self.wb = Workbook()
        # Remove default sheet (Workbook() always creates one).
        if self.wb.sheetnames:  # pragma: no branch
            del self.wb[self.wb.sheetnames[0]]

    def write(self) -> Workbook:
        """Write all table layouts and return the workbook."""
        # Sort layouts alphabetically by table code
        sorted_layouts = sorted(self.layouts, key=lambda lo: lo.table_code)

        # Identities are shared datapoints, so they are computed over
        # the whole workbook — one module for export_module().
        self.identities = _build_identity_index(sorted_layouts)

        entries: list[_IndexEntry] = []
        used_titles: set[str] = set()
        for layout in sorted_layouts:
            for active_sheets, sheet_id, sheet in _sheet_renderings(layout):
                title = _worksheet_title(
                    layout.table_code,
                    sheet,
                    used_titles,
                )
                self._write_table(layout, active_sheets, sheet_id, title)
                entries.append(
                    _IndexEntry(
                        title=title,
                        layout=layout,
                        sheet_label=sheet.label if sheet else "",
                    ),
                )

        # Add index sheet at the beginning
        self._write_index(entries)

        return self.wb

    def _write_index(self, entries: list[_IndexEntry]) -> None:
        """Write an index sheet with hyperlinks to each worksheet."""
        ws = self.wb.create_sheet(title="Index", index=0)

        # Title
        ws.cell(row=1, column=1, value="Table Index").font = Font(
            bold=True,
            size=14,
        )

        # A "Sheet" column is only meaningful when at least one table
        # was split into one worksheet per Z-axis sheet.
        has_sheets = any(e.sheet_label for e in entries)

        # Column headers
        headers = ["#", "Table Code", "Table Name"]
        if has_sheets:
            headers.append("Sheet")
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = _HEADER_FONT
            cell.fill = _GREY_FILL
            cell.border = _BORDER_ALL

        # Table entries with hyperlinks
        for i, entry in enumerate(entries):
            row = 4 + i
            ws.cell(row=row, column=1, value=i + 1).border = _BORDER_ALL

            code_cell = ws.cell(
                row=row, column=2, value=entry.layout.table_code
            )
            code_cell.border = _BORDER_ALL
            code_cell.font = Font(color="0000CC", underline="single")
            # An internal link needs a quoted location: titles carry
            # dots, spaces and parentheses, which Excel rejects
            # unquoted.
            code_cell.hyperlink = Hyperlink(
                ref=code_cell.coordinate,
                location=f"'{entry.title}'!A1",
            )

            name_cell = ws.cell(
                row=row, column=3, value=entry.layout.table_name
            )
            name_cell.border = _BORDER_ALL

            if has_sheets:
                sheet_cell = ws.cell(
                    row=row, column=4, value=entry.sheet_label
                )
                sheet_cell.border = _BORDER_ALL

        # Column widths
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 80
        if has_sheets:
            ws.column_dimensions["D"].width = 50

    def _write_table(  # noqa: C901
        self,
        layout: TableLayout,
        active_sheets: list[LayoutHeader],
        sheet_id: Optional[int],
        title: str,
    ) -> None:
        """Write one grid of a table layout to a worksheet.

        ``active_sheets`` are the Z-axis headers annotated on this
        worksheet and ``sheet_id`` the sheet its data cells are read
        from: one worksheet per Z sheet for sheet-scoped tables, a
        single worksheet carrying every Z header otherwise.
        """
        ws = self.wb.create_sheet(title=title)

        # True when this worksheet renders one specific Z sheet (rather
        # than a single grid whose Z headers are fixed context).
        per_sheet = bool(active_sheets) and (
            active_sheets[0].header_id == sheet_id
        )
        sheet_code = active_sheets[0].code if per_sheet else ""

        cfg = self.config
        code_row_offset = 1 if cfg.show_code_row else 0
        code_col_offset = 1 if cfg.show_code_column else 0

        # Layout geometry
        # Row 1: title
        # Row 2: Z-axis annotation label (if sheets) or empty
        # Row 3+: one row per annotated Z header, then a gap row
        # Next: "Columns" label + row annotation dimension headers
        # Next: column header depth levels (start + 1 + depth)
        # Then: column code row (if enabled)
        # Then: data rows

        # "Columns" row: row 4 without sheets, one row lower per
        # annotated Z header.
        col_header_start_row = 4 + len(active_sheets)
        col_label_rows = layout.max_col_depth + 1
        col_code_row = (
            col_header_start_row + 1 + col_label_rows
            if cfg.show_code_row
            else 0
        )
        data_start_row = (
            col_header_start_row + 1 + col_label_rows + code_row_offset
        )

        # Column geometry
        # Col A: "Rows" label (merged)
        # Col B: row labels
        # Col C: row codes (if enabled)
        # Cols after: data columns
        row_label_col = 2  # B
        row_code_col = 3 if cfg.show_code_column else 0
        data_start_col = (
            2 + code_col_offset + 1
        )  # after labels + optional codes

        # --- Build column position map ---
        # VBA logic: if the previous header was abstract, the current one
        # shares its position (abstract occupies same column as first child).
        # Otherwise, increment.
        col_positions: dict[int, int] = {}
        pos = 0
        prev_abstract = False
        for ch in layout.columns:
            if prev_abstract:
                col_positions[ch.header_id] = pos  # share with abstract
            else:
                pos += 1
                col_positions[ch.header_id] = pos
            prev_abstract = ch.is_abstract

        num_visible_cols = pos

        # --- Title ---
        title_text = f"{layout.table_code} - {layout.table_name}"
        title_cell = ws.cell(row=1, column=1, value=title_text)
        title_cell.font = _TITLE_FONT
        title_cell.fill = _GREY_FILL
        title_cell.border = _BORDER_ALL
        # Merge title across all columns. data_start_col is always
        # >= 3 (label cols + optional code col), so end_col is always
        # > 1 when there's any content; the guard is defensive.
        end_col = data_start_col + num_visible_cols - 1
        if end_col > 1:  # pragma: no branch
            ws.merge_cells(
                start_row=1,
                start_column=1,
                end_row=1,
                end_column=max(end_col, 1),
            )

        # --- Sheet header (Z-axis) ---
        # For tables with sheets:
        #   Row 2: annotation label (dimension name in data columns)
        #   Row 3+: "Sheet: {label}" (one worksheet per sheet) or
        #           "Sheet per {label}" (Z header as fixed context),
        #           plus the annotation value in the data columns
        #   Following row: empty gap
        if active_sheets:
            # Row 2: annotation label in data_start_col
            # Use "(member_code:domain_code) label" format for ATY dim
            for dm in active_sheets[0].categorisations:
                label_text = (
                    f"({dm.member_code}:{dm.domain_code}) {dm.member_label}"
                )
                ann_label_cell = ws.cell(
                    row=2, column=data_start_col, value=label_text
                )
                ann_label_cell.font = Font(bold=True, size=9)
                ann_label_cell.alignment = Alignment(horizontal="left")
                break  # only write for first (ATY) categorisation

            for sheet_idx, sh in enumerate(active_sheets):
                sheet_row = 3 + sheet_idx

                # Sheet header label + annotation value
                prefix = "Sheet: " if per_sheet else "Sheet per "
                sc = ws.cell(
                    row=sheet_row,
                    column=row_label_col,
                    value=f"{prefix}{sh.label}",
                )
                sc.font = _HEADER_FONT
                sc.fill = _GREY_FILL
                sc.border = _BORDER_ALL
                sc.alignment = Alignment(
                    horizontal="center", vertical="center"
                )

                # Annotation value: SubCategory info
                if sh.subcategory_code and sh.subcategory_cat_code:
                    cat = sh.subcategory_cat_code
                    sc_code = sh.subcategory_code
                    sc_desc = sh.subcategory_description
                    ann_val = f"({cat}:{sc_code}({cat}))  ({sc_desc})"
                    if sh.is_key:
                        ann_val += " <Key value> "
                    ann_val_cell = ws.cell(
                        row=sheet_row, column=data_start_col, value=ann_val
                    )
                    ann_val_cell.font = Font(size=9)
                    ann_val_cell.alignment = Alignment(horizontal="left")

                if cfg.add_header_comments and sh.categorisations:
                    tooltip = _format_categorisations(sh.categorisations)
                    sc.comment = Comment(
                        tooltip, "dpmcore", width=400, height=150
                    )

        # --- "Columns" label ---
        col_cell = ws.cell(
            row=col_header_start_row, column=data_start_col, value="Columns"
        )
        col_cell.font = _HEADER_FONT
        col_cell.fill = _GREY_FILL
        col_cell.border = Border(
            left=_THIN, right=_THIN, top=_THIN, bottom=Side(style=None)
        )
        col_cell.alignment = Alignment(horizontal="center")
        if num_visible_cols > 1:
            ws.merge_cells(
                start_row=col_header_start_row,
                start_column=data_start_col,
                end_row=col_header_start_row,
                end_column=data_start_col + num_visible_cols - 1,
            )

        # --- Column headers ---
        for ch in layout.columns:
            col_offset = col_positions[ch.header_id]
            if col_offset == 0:  # pragma: no cover
                # Defensive: pos starts at 0 and is always incremented
                # before assignment, so col_positions values are >= 1.
                continue

            excel_col = data_start_col - 1 + col_offset

            # Column label at (col_header_start_row + 1 + depth)
            label_row = col_header_start_row + 1 + ch.depth
            label_text = ch.label
            if not cfg.show_code_row and not ch.is_abstract:
                label_text = f"{ch.code} {ch.label}"
            if ch.is_abstract and cfg.show_abstract_header_codes:
                label_text = f"{ch.code} {ch.label}"

            cell = ws.cell(row=label_row, column=excel_col, value=label_text)
            cell.fill = _GREY_FILL
            cell.border = _BORDER_ALL
            cell.alignment = Alignment(
                horizontal="center",
                vertical="top",
                wrap_text=True,
            )
            cell.font = _HEADER_FONT

            # Column code row
            if cfg.show_code_row and not ch.is_abstract and col_code_row:
                code_cell = ws.cell(row=col_code_row, column=excel_col)
                code_cell.value = ch.code
                code_cell.number_format = "@"  # text format
                code_cell.fill = _GREY_FILL
                code_cell.border = _BORDER_ALL
                code_cell.alignment = Alignment(horizontal="center")

            # Header tooltip
            if cfg.add_header_comments and ch.categorisations:
                tooltip = _format_categorisations(ch.categorisations)
                cell.comment = Comment(
                    tooltip, "dpmcore", width=400, height=150
                )

        # Grey-fill the entire column header grid (including empty cells)
        if num_visible_cols > 0 and layout.max_col_depth >= 0:
            header_top = col_header_start_row
            header_bottom = col_header_start_row + 1 + layout.max_col_depth
            if cfg.show_code_row:
                header_bottom += 1
            for r in range(header_top, header_bottom + 1):
                for c_off in range(1, num_visible_cols + 1):
                    ec = data_start_col - 1 + c_off
                    cell = ws.cell(row=r, column=ec)
                    if cell.fill.fill_type is None:  # unfilled
                        cell.fill = _GREY_FILL

        # Merge column headers: vertical borders for leaves,
        # horizontal merges for parents.
        _merge_column_headers(
            ws,
            layout.columns,
            col_positions,
            col_header_start_row + 1,
            data_start_col,
            layout.max_col_depth,
        )

        # --- Row headers ---
        # "Rows" label
        rows_label_cell = ws.cell(
            row=data_start_row,
            column=1,
            value="Rows",
        )
        rows_label_cell.font = _HEADER_FONT
        rows_label_cell.fill = _GREY_FILL
        rows_label_cell.border = _BORDER_ALL
        rows_label_cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            text_rotation=90,
        )

        # Merge "Rows" label vertically
        if len(layout.rows) > 1:
            ws.merge_cells(
                start_row=data_start_row,
                start_column=1,
                end_row=data_start_row + len(layout.rows) - 1,
                end_column=1,
            )

        # Row header entries
        row_positions: dict[int, int] = {}
        for i, rh in enumerate(layout.rows):
            excel_row = data_start_row + i
            row_positions[rh.header_id] = excel_row

            # Row label with indentation
            if rh.is_abstract or not cfg.show_code_column:
                label_text = f"{rh.code} {rh.label}"
            else:
                label_text = rh.label

            indent = max(0, rh.depth) * 2
            label_cell = ws.cell(
                row=excel_row,
                column=row_label_col,
                value=label_text,
            )
            label_cell.fill = _GREY_FILL
            label_cell.border = _BORDER_ALL
            label_cell.alignment = Alignment(
                horizontal="left",
                vertical="center",
                indent=indent,
                wrap_text=True,
            )

            if rh.is_abstract:
                label_cell.font = Font(bold=True, size=10)
                # Merge abstract row across all data columns
                if num_visible_cols > 0:
                    merge_end = data_start_col + num_visible_cols - 1
                    ws.merge_cells(
                        start_row=excel_row,
                        start_column=row_label_col,
                        end_row=excel_row,
                        end_column=merge_end,
                    )
            else:
                label_cell.font = Font(size=10)

            # Row code column
            if cfg.show_code_column and not rh.is_abstract:
                code_cell = ws.cell(row=excel_row, column=row_code_col)
                code_cell.value = rh.code
                code_cell.number_format = "@"  # text format
                code_cell.border = _BORDER_ALL

            # Header tooltip
            if cfg.add_header_comments and rh.categorisations:
                tooltip = _format_categorisations(rh.categorisations)
                label_cell.comment = Comment(
                    tooltip,
                    "dpmcore",
                    width=400,
                    height=150,
                )

        # --- Open-row tables ---
        if layout.is_open_row:
            self._write_open_row_data(
                ws,
                layout,
                data_start_row,
                data_start_col,
                col_positions,
                row_label_col,
                cfg,
                sheet_id,
                sheet_code,
            )

        # --- Data cells ---
        for rh in layout.rows:
            if rh.is_abstract:
                continue
            excel_row = row_positions[rh.header_id]

            for ch in layout.columns:
                if ch.is_abstract:
                    continue
                col_offset = col_positions[ch.header_id]
                excel_col = data_start_col - 1 + col_offset

                # Try to find cell data
                cell_data = layout.cells.get(
                    (rh.header_id, ch.header_id, sheet_id),
                )

                cell = ws.cell(row=excel_row, column=excel_col)
                cell.font = _DATA_FONT
                cell.border = _BORDER_ALL

                if cell_data is not None and cell_data.is_void:
                    cell.fill = _VOID_FILL
                elif cell_data is None or cell_data.is_excluded:
                    cell.fill = _EXCLUDED_FILL
                elif cell_data.variable_vid:
                    # Build cell content: VariableID + data type + sign
                    display_id = (
                        cell_data.variable_id or cell_data.variable_vid
                    )
                    cell_lines = [str(display_id)]
                    if cell_data.data_type_code:
                        if (
                            cell_data.data_type_code == "e"
                            and cell_data.domain_label
                        ):
                            cell_lines.append(f"[{cell_data.domain_label}]")
                        elif cell_data.data_type_code in _DATA_TYPE_SYMBOLS:
                            cell_lines.append(
                                _DATA_TYPE_SYMBOLS[cell_data.data_type_code]
                            )
                    # Only the sign stored on the cell is shown: a
                    # NULL sign is "No Sign" in DPM Studio and must not
                    # be reported as positive (DRR-1967, DRR-1970).
                    if cell_data.sign:
                        cell_lines.append(cell_data.sign)
                    cell.value = "\n".join(cell_lines)
                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center",
                        wrap_text=True,
                    )

                    # Identities: highlight datapoints reported by more
                    # than one cell of the workbook.
                    location = _cell_location(
                        layout.table_code,
                        rh.code,
                        ch.code,
                        sheet_code,
                    )
                    others = self._other_identity_locations(
                        cell_data.variable_vid,
                        location,
                    )
                    if others:
                        cell.fill = _IDENTITY_FILL

                    # Cell tooltip
                    if cfg.add_cell_comments:
                        cell.comment = Comment(
                            _cell_tooltip(cell_data, others),
                            "dpmcore",
                            width=400,
                            height=200,
                        )

        # --- Annotations ---
        if cfg.annotate:
            self._write_annotations(
                ws,
                layout,
                data_start_row,
                data_start_col,
                row_positions,
                col_positions,
                num_visible_cols,
                row_label_col,
                col_header_start_row,
                sheet_id,
            )

        # --- Outline groups ---
        _apply_row_groups(ws, layout.rows, row_positions)
        _apply_col_groups(
            ws,
            layout.columns,
            col_positions,
            data_start_col,
        )

        # --- Freeze panes ---
        freeze_ref = f"{get_column_letter(data_start_col)}{data_start_row}"
        ws.freeze_panes = freeze_ref

        # --- Column widths ---
        ws.column_dimensions["A"].width = 5
        ws.column_dimensions[get_column_letter(row_label_col)].width = 40
        if cfg.show_code_column:
            ws.column_dimensions[get_column_letter(row_code_col)].width = 8
        for col_offset in range(1, num_visible_cols + 1):
            col_letter = get_column_letter(data_start_col - 1 + col_offset)
            ws.column_dimensions[col_letter].width = 18

    def _write_open_row_data(  # noqa: C901
        self,
        ws: Any,
        layout: TableLayout,
        data_start_row: int,
        data_start_col: int,
        col_positions: dict[int, int],
        row_label_col: int,
        cfg: ExportConfig,
        sheet_id: Optional[int],
        sheet_code: str,
    ) -> None:
        """Write the single 'Open Rows' data row for open-row tables."""
        open_rows_cell = ws.cell(
            row=data_start_row,
            column=row_label_col,
            value="Open Rows",
        )
        open_rows_cell.fill = _GREY_FILL
        open_rows_cell.border = _BORDER_ALL
        open_rows_cell.alignment = Alignment(
            horizontal="left", vertical="center"
        )
        open_rows_cell.font = _HEADER_FONT

        for ch in layout.columns:
            if ch.is_abstract:
                continue
            col_offset = col_positions[ch.header_id]
            excel_col = data_start_col - 1 + col_offset

            cell = ws.cell(row=data_start_row, column=excel_col)
            cell.font = _DATA_FONT
            cell.border = _BORDER_ALL

            if ch.is_key and ch.key_variable_id:
                cell.fill = _KEY_FILL
                if ch.key_data_type_code == "e":
                    type_symbol = f"[{ch.key_property_name}]"
                else:
                    type_symbol = _DATA_TYPE_SYMBOLS.get(
                        ch.key_data_type_code,
                        ch.key_data_type_code,
                    )
                lines = [str(ch.key_variable_id), type_symbol]
                if ch.key_property_name:
                    lines.append(f"<{ch.key_property_name}>")
                cell.value = "\n".join(lines)
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )
            else:
                cell_data = layout.cells.get((None, ch.header_id, sheet_id))
                if (
                    cell_data is not None
                    and not cell_data.is_excluded
                    and cell_data.variable_vid
                ):
                    display_id = (
                        cell_data.variable_id or cell_data.variable_vid
                    )
                    cell_lines = [str(display_id)]
                    if cell_data.data_type_code:
                        if (
                            cell_data.data_type_code == "e"
                            and cell_data.domain_label
                        ):
                            cell_lines.append(f"[{cell_data.domain_label}]")
                        elif cell_data.data_type_code in _DATA_TYPE_SYMBOLS:
                            cell_lines.append(
                                _DATA_TYPE_SYMBOLS[cell_data.data_type_code]
                            )
                    if cell_data.sign:
                        cell_lines.append(cell_data.sign)
                    cell.value = "\n".join(cell_lines)
                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center",
                        wrap_text=True,
                    )

                    location = _cell_location(
                        layout.table_code,
                        "",
                        ch.code,
                        sheet_code,
                    )
                    others = self._other_identity_locations(
                        cell_data.variable_vid,
                        location,
                    )
                    if others:
                        cell.fill = _IDENTITY_FILL
                    if cfg.add_cell_comments:
                        cell.comment = Comment(
                            _cell_tooltip(cell_data, others),
                            "dpmcore",
                            width=400,
                            height=200,
                        )
                elif cell_data is not None and cell_data.is_void:
                    cell.fill = _VOID_FILL
                else:
                    cell.fill = _EXCLUDED_FILL

    def _other_identity_locations(
        self,
        variable_vid: int,
        location: str,
    ) -> list[str]:
        """Other cells of the workbook reporting the same datapoint."""
        return [
            loc
            for loc in self.identities.get(variable_vid, [])
            if loc != location
        ]

    def _write_annotations(  # noqa: C901
        self,
        ws: Any,
        layout: TableLayout,
        data_start_row: int,
        data_start_col: int,
        row_positions: dict[int, int],
        col_positions: dict[int, int],
        num_visible_cols: int,
        row_label_col: int,
        col_header_start_row: int,
        sheet_id: Optional[int] = None,
    ) -> None:
        """Write dimensional annotations around the data grid."""
        col_dims = _collect_axis_dimensions(layout.columns)

        # Some dimensions only appear in cell dp_categorisations but
        # not in any row header context. Add those as supplemental row
        # annotation dimensions, but only if they are NOT already shown
        # as column annotation dimensions (dimensions that vary by
        # column are column-level, not row-level).
        col_dim_set = set(col_dims)
        row_dp_cats: dict[int, list[DimensionMember]] = {}
        for cell_key, cell_data in layout.cells.items():
            row_id, _col_id, cell_sheet_id = cell_key
            if row_id is None or cell_sheet_id != sheet_id:
                continue
            if cell_data.dp_categorisations:
                extra = [
                    dm
                    for dm in cell_data.dp_categorisations
                    if _dim_display_label(dm) not in col_dim_set
                ]
                if extra:
                    row_dp_cats.setdefault(row_id, []).extend(extra)

        row_dims = _collect_axis_dimensions(
            layout.rows, supplemental_cats=row_dp_cats
        )

        # Assign colors consistently across both axes (keyed by display label)
        all_dim_labels = list(dict.fromkeys(col_dims + row_dims))
        dim_color_map: dict[str, str] = {
            label: _DIM_COLORS[i % len(_DIM_COLORS)]
            for i, label in enumerate(all_dim_labels)
        }

        # --- Column annotations (below the data grid, no gap row) ---
        if col_dims:
            # Open-row tables have no layout.rows but one
            # "Open Rows" data row.
            ann_start_row = data_start_row + (
                1 if layout.is_open_row else len(layout.rows)
            )

            for dim_idx, dim_label in enumerate(col_dims):
                ann_row = ann_start_row + dim_idx
                color = dim_color_map[dim_label]

                # Dimension label in row_label_col (col B)
                label_cell = ws.cell(
                    row=ann_row,
                    column=row_label_col,
                    value=dim_label,
                )
                label_cell.font = Font(bold=True, size=9, color=color)
                label_cell.alignment = Alignment(horizontal="right")

                # Member values per column
                for ch in layout.columns:
                    if ch.is_abstract:
                        continue
                    col_offset = col_positions[ch.header_id]
                    excel_col = data_start_col - 1 + col_offset

                    member = _find_member_by_label(
                        ch.categorisations, dim_label
                    )
                    if member:
                        display = _member_display_label(member)
                        if ch.subcategory_code and ch.subcategory_cat_code:
                            display += (
                                f"[{ch.subcategory_cat_code}:"
                                f"{ch.subcategory_code}]"
                            )
                    elif ch.is_key:
                        key_dm = _find_member_by_label(
                            ch.key_categorisations, dim_label
                        )
                        if key_dm:
                            display = _key_member_display(key_dm, ch)
                            # Use key_dm as sentinel for the write below.
                            member = key_dm
                    if member:
                        mc = ws.cell(
                            row=ann_row, column=excel_col, value=display
                        )
                        mc.font = Font(size=9, color=color)
                        mc.alignment = Alignment(
                            horizontal="center",
                            wrap_text=True,
                        )

        # --- Row annotations (right of the data grid, no gap column) ---
        if row_dims:
            ann_start_col = data_start_col + num_visible_cols
            # Span: col_header_start_row + max_col_depth to +1
            ann_header_row = col_header_start_row + layout.max_col_depth

            for dim_idx, dim_label in enumerate(row_dims):
                ann_col = ann_start_col + dim_idx
                color = dim_color_map[dim_label]

                # Dimension label at ann_header_row, merged with next row
                label_cell = ws.cell(
                    row=ann_header_row,
                    column=ann_col,
                    value=dim_label,
                )
                label_cell.font = Font(bold=True, size=9, color=color)
                label_cell.alignment = Alignment(
                    horizontal="left",
                    vertical="bottom",
                    wrap_text=True,
                )
                ws.merge_cells(
                    start_row=ann_header_row,
                    start_column=ann_col,
                    end_row=ann_header_row + 1,
                    end_column=ann_col,
                )

                # Member values per row
                for rh in layout.rows:
                    if rh.is_abstract:
                        continue
                    excel_row = row_positions[rh.header_id]

                    member = _find_member_by_label(
                        rh.categorisations, dim_label
                    )
                    if member is None and rh.header_id in row_dp_cats:
                        member = _find_member_by_label(
                            row_dp_cats[rh.header_id], dim_label
                        )
                    if member:
                        display = _member_display_label(member)
                        if rh.subcategory_code and rh.subcategory_cat_code:
                            display += (
                                f"[{rh.subcategory_cat_code}:"
                                f"{rh.subcategory_code}]"
                            )
                        mc = ws.cell(
                            row=excel_row, column=ann_col, value=display
                        )
                        mc.font = Font(size=9, color=color)
                        mc.alignment = Alignment(horizontal="left")

                # Set annotation column width
                ws.column_dimensions[get_column_letter(ann_col)].width = 25


def _sheet_renderings(
    layout: TableLayout,
) -> list[tuple[list[LayoutHeader], Optional[int], Optional[LayoutHeader]]]:
    """Split a layout into the worksheets it needs.

    Cells of a Z-axis table are keyed by the sheet they belong to, so
    such a table is rendered as one worksheet per Z sheet. Tables whose
    cells carry no sheet get a single worksheet on which every Z header
    is annotated as fixed context.

    Returns ``(annotated Z headers, sheet id for cell lookup, sheet)``
    per worksheet; ``sheet`` is None for a single-worksheet table.
    """
    cell_sheet_ids = {key[2] for key in layout.cells if key[2] is not None}
    if layout.sheets and cell_sheet_ids:
        per_sheet: list[
            tuple[list[LayoutHeader], Optional[int], Optional[LayoutHeader]]
        ] = [
            ([sh], sh.header_id, sh)
            for sh in layout.sheets
            if sh.header_id in cell_sheet_ids
        ]
        if per_sheet:
            return per_sheet

    # Single grid: the sheet id the cells were keyed with (usually None).
    default_sheet_id = None
    for key in layout.cells:
        default_sheet_id = key[2]
        break
    return [(layout.sheets, default_sheet_id, None)]


def _worksheet_title(
    table_code: str,
    sheet: Optional[LayoutHeader],
    used: set[str],
) -> str:
    """Build a unique worksheet title of at most 31 characters."""
    suffix = f" ({sheet.code or sheet.label})" if sheet is not None else ""
    title = table_code[: _MAX_SHEET_TITLE - len(suffix)] + suffix
    if title in used:
        # Distinct sheets can collide once truncated; keep them apart.
        counter = 2
        while True:
            tag = f"~{counter}"
            candidate = title[: _MAX_SHEET_TITLE - len(tag)] + tag
            if candidate not in used:
                title = candidate
                break
            counter += 1
    used.add(title)
    return title


def _cell_location(
    table_code: str,
    row_code: str,
    col_code: str,
    sheet_code: str,
) -> str:
    """Human-readable cell address used to report identities."""
    parts = [table_code]
    if sheet_code:
        parts.append(f"s{sheet_code}")
    if row_code:
        parts.append(f"r{row_code}")
    if col_code:
        parts.append(f"c{col_code}")
    return " ".join(parts)


def _build_identity_index(layouts: list[TableLayout]) -> dict[int, list[str]]:
    """Map VariableVID to the cells reporting it, for shared datapoints.

    Cells sharing a datapoint are "identities": the same figure is
    reported in more than one place. Only variables used by at least
    two cells are returned.
    """
    locations: dict[int, list[str]] = {}
    for layout in layouts:
        codes: dict[int, str] = {
            h.header_id: h.code
            for h in layout.rows + layout.columns + layout.sheets
        }
        for (row_id, col_id, sheet_id), cd in layout.cells.items():
            if not cd.variable_vid or cd.is_excluded or cd.is_void:
                continue
            locations.setdefault(cd.variable_vid, []).append(
                _cell_location(
                    layout.table_code,
                    codes.get(row_id, "") if row_id is not None else "",
                    codes.get(col_id, ""),
                    codes.get(sheet_id, "") if sheet_id is not None else "",
                ),
            )
    return {vvid: locs for vvid, locs in locations.items() if len(locs) > 1}


def _cell_tooltip(cell_data: CellData, identities: list[str]) -> str:
    """Tooltip for a data cell: variable ids, identities, dimensions."""
    lines = []
    if cell_data.variable_id:
        lines.append(f"VariableID = {cell_data.variable_id}")
    lines.append(f"VariableVID = {cell_data.variable_vid}")
    if identities:
        shown = identities[:_MAX_IDENTITIES_IN_TOOLTIP]
        lines.append("")
        lines.append("Identity - same datapoint also reported in:")
        lines.extend(f"  {loc}" for loc in shown)
        hidden = len(identities) - len(shown)
        if hidden:
            lines.append(f"  ... and {hidden} more")
    if cell_data.dp_categorisations:
        lines.append("")
        lines.append(_format_categorisations(cell_data.dp_categorisations))
    return "\n".join(lines)


def _collect_axis_dimensions(
    headers: list[LayoutHeader],
    supplemental_cats: Optional[dict[int, list[DimensionMember]]] = None,
) -> list[str]:
    """Collect dimensions that have at least one value on this axis.

    Returns a sorted list of dimension display labels, deduplicated by label
    (so "Main Property" / ATY appears once). ATY dimensions sort first.

    supplemental_cats: optional dict mapping header_id ->
    list[DimensionMember] from dp_categorisations, to pick up
    dimensions not in header categorisations.
    """
    # key=label -> (first_property_id, dimension_code) for sort ordering
    seen: dict[str, tuple[int, str]] = {}
    for h in headers:
        if h.is_abstract:
            continue
        all_cats = list(h.categorisations)
        if supplemental_cats and h.header_id in supplemental_cats:
            all_cats.extend(supplemental_cats[h.header_id])
        if h.is_key:
            all_cats.extend(h.key_categorisations)
        for dm in all_cats:
            label = _dim_display_label(dm)
            if label not in seen:
                seen[label] = (dm.property_id, dm.dimension_code or "")

    def _sort_key(label: str) -> tuple[int, int]:
        prop_id, dim_code = seen[label]
        # ATY dimensions (Main Property) sort before all others
        return (0 if dim_code == "ATY" else 1, prop_id)

    return sorted(seen.keys(), key=_sort_key)


def _dim_display_label(dm: DimensionMember) -> str:
    """Format a dimension header label: ``(DimCode:DomainCode) Label``."""
    if dm.dimension_code == "ATY" or not dm.dimension_code:
        return dm.dimension_label
    return f"({dm.dimension_code}:{dm.domain_code}) {dm.dimension_label}"


def _member_display_label(dm: DimensionMember) -> str:
    """Format a member annotation: ``(DomainCode:MemberCode) Label``."""
    if dm.dimension_code == "ATY":
        if dm.member_code:
            return f"({dm.member_code}) {dm.member_label}"
        return dm.member_label
    if dm.domain_code and dm.member_code:
        return f"({dm.domain_code}:{dm.member_code}) {dm.member_label}"
    return dm.member_label


def _key_member_display(dm: DimensionMember, ch: LayoutHeader) -> str:
    """Format the member value for a key column annotation as '<Key value>'."""
    if ch.subcategory_code and ch.subcategory_cat_code:
        return (
            f"({dm.domain_code}:"
            f"{ch.subcategory_code}({ch.subcategory_cat_code})) "
            f"<Key value>[{ch.subcategory_cat_code}:"
            f"{ch.subcategory_code}] "
            f"{ch.subcategory_description}"
        )
    return f"({dm.domain_code}:) <Key value>"


def _find_member_by_label(
    cats: list[DimensionMember],
    dim_display_label: str,
) -> Optional[DimensionMember]:
    """Find a DimensionMember matching a given dimension display label."""
    for dm in cats:
        if _dim_display_label(dm) == dim_display_label:
            return dm
    return None


def _format_categorisations(cats: list[DimensionMember]) -> str:
    """Format categorisations as tooltip text: Dimension = Member."""
    lines = [f"{dm.dimension_label}  =  {dm.member_label}" for dm in cats]
    return "\n".join(lines)


def _merge_column_headers(  # noqa: C901
    ws: Any,
    columns: list[LayoutHeader],
    col_positions: dict[int, int],
    col_header_start_row: int,
    data_start_col: int,
    max_depth: int,
) -> None:
    """Apply column header merging and border patterns.

    col_header_start_row is the FIRST column header row (depth-0 row).

    - Leaf columns at depth < max_depth: remove bottom border, add empty
      cells below with vertical border pattern (no merge).
    - Non-leaf parent columns (abstract or not): horizontal merge across
      their descendant span; add empty cells below with border pattern.
    """
    parent_ids: set[int] = set()
    for ch in columns:
        if ch.parent_header_id is not None:
            parent_ids.add(ch.parent_header_id)

    by_id: dict[int, LayoutHeader] = {ch.header_id: ch for ch in columns}

    # Pass 1: border pattern for non-abstract columns at depth < max_depth
    for ch in columns:
        if ch.is_abstract:
            continue
        col_offset = col_positions.get(ch.header_id, 0)
        if col_offset == 0:
            continue
        excel_col = data_start_col - 1 + col_offset
        label_row = col_header_start_row + ch.depth
        last_depth_row = col_header_start_row + max_depth

        if ch.depth < max_depth:
            # Remove bottom border so the empty cells below visually connect
            # with the label. Non-abstract parents need this too: descendants
            # sit at other column positions, so the parent's own column has
            # only empty cells below the label. The descendant columns get
            # their separator from the next-depth label's own top border.
            cell = ws.cell(row=label_row, column=excel_col)
            cell.border = Border(
                left=_THIN,
                top=_THIN,
                right=_THIN,
                bottom=Side(style=None),
            )

            # Empty cells below: intermediate=L+R, last=L+R+B (no top)
            for r in range(label_row + 1, last_depth_row + 1):
                empty_cell = ws.cell(row=r, column=excel_col)
                empty_cell.fill = _GREY_FILL
                if r < last_depth_row:
                    empty_cell.border = Border(left=_THIN, right=_THIN)
                else:
                    empty_cell.border = Border(
                        left=_THIN,
                        right=_THIN,
                        bottom=_THIN,
                    )

    # Pass 2: horizontal merge for all parent headers
    # (abstract and non-abstract).
    for ch in columns:
        if ch.header_id not in parent_ids:
            continue
        col_offset = col_positions.get(ch.header_id, 0)
        if col_offset == 0:
            continue

        # Find positions of all non-abstract descendants
        desc_positions: list[int] = []
        for other in columns:
            if other.header_id == ch.header_id:
                continue
            if not other.is_abstract and _is_descendant(
                other, ch.header_id, by_id
            ):
                p = col_positions.get(other.header_id, 0)
                if p > 0:
                    desc_positions.append(p)

        if desc_positions:
            min_pos = col_offset  # parent's own column
            max_pos = max(desc_positions)
            start_col = data_start_col - 1 + min_pos
            end_col = data_start_col - 1 + max_pos
            label_row = col_header_start_row + ch.depth

            if end_col > start_col:
                ws.merge_cells(
                    start_row=label_row,
                    start_column=start_col,
                    end_row=label_row,
                    end_column=end_col,
                )


def _is_descendant(
    header: LayoutHeader,
    ancestor_id: int,
    by_id: dict[int, LayoutHeader],
) -> bool:
    """Check if header is a descendant of ancestor_id."""
    current = header
    visited: set[int] = set()
    while current.parent_header_id is not None:
        if current.parent_header_id in visited:
            break
        visited.add(current.parent_header_id)
        if current.parent_header_id == ancestor_id:
            return True
        current = by_id.get(current.parent_header_id)  # type: ignore[assignment]
        if current is None:
            break
    return False


def _apply_row_groups(
    ws: Any,
    rows: list[LayoutHeader],
    row_positions: dict[int, int],
) -> None:
    """Apply Excel outline grouping to hierarchical rows."""
    for rh in rows:
        if rh.depth > 0:
            excel_row = row_positions.get(rh.header_id)
            if excel_row:
                ws.row_dimensions[excel_row].outline_level = min(rh.depth, 7)


def _apply_col_groups(
    ws: Any,
    columns: list[LayoutHeader],
    col_positions: dict[int, int],
    data_start_col: int,
) -> None:
    """Apply Excel outline grouping to hierarchical columns."""
    for ch in columns:
        if ch.depth > 0 and not ch.is_abstract:
            col_offset = col_positions.get(ch.header_id, 0)
            if col_offset > 0:
                excel_col = data_start_col - 1 + col_offset
                col_letter = get_column_letter(excel_col)
                ws.column_dimensions[col_letter].outline_level = min(
                    ch.depth,
                    7,
                )
