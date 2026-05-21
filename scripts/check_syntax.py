"""Check DPM-XL expression syntax for all rules in a validations xlsx."""

import argparse
import sys
from pathlib import Path
from typing import Any

import openpyxl

from dpmcore.services.syntax import SyntaxService

_DEFAULT_XLSX = (
    Path(__file__).resolve().parents[1]
    / "tests"
    / "fixtures"
    / "validations_export.xlsx"
)


def _check_rows(
    rows: list[dict[str, Any]],
) -> tuple[list[tuple[str, str, str, str]], int]:
    svc = SyntaxService()
    items = [
        (str(row.get("Code") or ""), col, str(row.get(col)).strip())
        for row in rows
        for col in ("Expression", "Precondition")
        if row.get(col) is not None and str(row.get(col)).strip()
    ]
    total = len(items)
    width = len(str(total))
    failures: list[tuple[str, str, str, str]] = []
    for i, (code, col, expr) in enumerate(items, start=1):
        result = svc.validate(expr)
        if result.is_valid:
            print(f"[{i:{width}}/{total}] {code} | {col} | PASS")
        else:
            error = result.error_message or ""
            print(f"[{i:{width}}/{total}] {code} | {col} | FAIL: {error}")
            print(f"  Expression: {expr}")
            failures.append((code, col, expr, error))
    return failures, total


def main() -> int:
    """Run syntax checks and print results to stdout."""
    parser = argparse.ArgumentParser(
        description="Check DPM-XL expression syntax in a validations xlsx.",
    )
    parser.add_argument(
        "--xlsx",
        default=str(_DEFAULT_XLSX),
        help=f"Path to the validations xlsx (default: {_DEFAULT_XLSX})",
    )
    args = parser.parse_args()

    print("Syntax check")
    print(f"  xlsx: {args.xlsx}")
    print()

    wb = openpyxl.load_workbook(args.xlsx)
    ws = wb["Validations"]
    headers = [cell.value for cell in ws[1]]
    rows: list[dict[str, Any]] = [
        dict(zip(headers, row, strict=False))
        for row in ws.iter_rows(min_row=2, values_only=True)
    ]

    failures, total = _check_rows(rows)

    passed = total - len(failures)
    failed = len(failures)
    print(f"\n{total} expressions checked — {passed} passed, {failed} failed")

    if failures:
        width_f = len(str(len(failures)))
        print("\nFailed expressions:")
        for j, (code, col, expr, error) in enumerate(failures, 1):
            prefix = f"  [{j:{width_f}}/{len(failures)}] {code} | {col}"
            print(f"{prefix} | FAIL: {error}")
            print(f"    Expression: {expr}")

    return 1 if failures else 0


if __name__ == "__main__":
    sys.exit(main())
