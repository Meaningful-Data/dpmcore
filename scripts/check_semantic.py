"""Check DPM-XL expression semantics for all rules in a validations xlsx."""

import argparse
import sys
from pathlib import Path
from typing import Any

import openpyxl
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from dpmcore.services.semantic import SemanticService

_DEFAULT_XLSX = (
    Path(__file__).resolve().parents[1]
    / "tests"
    / "fixtures"
    / "validations_export.xlsx"
)
_DEFAULT_DB = (
    Path(__file__).resolve().parents[1] / "tests" / "fixtures" / "test_data.db"
)


def _check_rows(
    rows: list[dict[str, Any]],
    svc: SemanticService,
) -> tuple[list[tuple[str, str, str, str, str]], int]:
    items = [
        (
            str(row.get("Code") or ""),
            col,
            str(row.get("StartRelease") or "").strip(),
            str(row.get(col)).strip(),
        )
        for row in rows
        for col in ("Expression", "Precondition")
        if row.get(col) is not None and str(row.get(col)).strip()
    ]
    total = len(items)
    width = len(str(total))
    failures: list[tuple[str, str, str, str, str]] = []
    for i, (code, col, release, expr) in enumerate(items, start=1):
        result = svc.validate(expr, release_code=release or None)
        prefix = f"[{i:{width}}/{total}] {code} | {col} | release {release}"
        if result.is_valid:
            print(f"{prefix} | PASS")
        else:
            error = result.error_message or ""
            print(f"{prefix} | FAIL: {error}")
            print(f"  Expression: {expr}")
            failures.append((code, col, release, expr, error))
    return failures, total


def main() -> int:
    """Run semantic checks and print results to stdout."""
    parser = argparse.ArgumentParser(
        description="Check DPM-XL expression semantics in a validations xlsx.",
    )
    parser.add_argument(
        "--xlsx",
        default=str(_DEFAULT_XLSX),
        help=f"Path to the validations xlsx (default: {_DEFAULT_XLSX})",
    )
    parser.add_argument(
        "--db",
        default=str(_DEFAULT_DB),
        help=f"Path to the SQLite DPM database (default: {_DEFAULT_DB})",
    )
    args = parser.parse_args()

    print("Semantic check")
    print(f"  xlsx: {args.xlsx}")
    print(f"  db:   {args.db}")
    print()

    engine = create_engine(f"sqlite:///{args.db}")
    Session = sessionmaker(bind=engine)
    session = Session()

    try:
        wb = openpyxl.load_workbook(args.xlsx)
        ws = wb["Validations"]
        headers = [cell.value for cell in ws[1]]
        rows: list[dict[str, Any]] = [
            dict(zip(headers, row, strict=False))
            for row in ws.iter_rows(min_row=2, values_only=True)
        ]

        svc = SemanticService(session)
        failures, total = _check_rows(rows, svc)

        passed = total - len(failures)
        failed = len(failures)
        print(
            f"\n{total} expressions checked — {passed} passed, {failed} failed"
        )

        if failures:
            width_f = len(str(len(failures)))
            print("\nFailed expressions:")
            for j, (code, col, release, expr, error) in enumerate(failures, 1):
                print(
                    f"  [{j:{width_f}}/{len(failures)}]"
                    f" {code} | {col} | release {release} | FAIL: {error}"
                )
                print(f"    Expression: {expr}")

        return 1 if failures else 0
    finally:
        session.close()
        engine.dispose()


if __name__ == "__main__":
    sys.exit(main())
